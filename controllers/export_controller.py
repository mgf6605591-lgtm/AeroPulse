# controllers/export_controller.py
"""Сборка книги XLSX: числа числами, разметка — как на экране.

Слой не знает ни о Qt, ни о том, откуда взялись значения (ARCH-2). Прежде
экспортёр сам читал `QTableView` и сам показывал `QMessageBox` — то есть был
кодом интерфейса, живущим в слое, который по замыслу интерфейса не касается.
Позвать его из теста или из командной строки было нельзя.

Чтение модели и окна сообщений переехали в [forms/table_export.py](forms/table_export.py);
сюда приходят готовые заголовки и строки значений.

Часть ячеек свода — суммы соседних, и они уходят в книгу формулами: правила
приносит `formulas` ([controllers/reports/formulas.py](controllers/reports/formulas.py)),
а сверяются они здесь, где есть сами значения.
"""
import math
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from controllers.export_header import ExportHeader
from controllers.reports.formulas import CellRef, FormulaMap


class ExportController:
    """Записывает книгу по готовым заголовкам и строкам значений."""

    # Формат разрядности. Запятая и точка в коде формата — не символы, а
    # обозначения разделителей: Excel подставляет принятые в своей локали,
    # то есть в русской покажет «1 234,57».
    INT_FORMAT = "#,##0"
    DECIMAL_FORMAT = "#,##0.00"

    # Ведущие символы, с которых Excel начинает разбор ячейки как формулы.
    FORMULA_STARTERS = ("=", "+", "-", "@")

    # Насколько сумма слагаемых может разойтись с самой ячейкой, чтобы формула
    # считалась сверенной. Допуск покрывает только двоичное сложение, но не
    # расхождение отчёта: значения хранятся десятичными с той точностью, с какой
    # пришли (db/models/types.py), и наименьшее осмысленное расхождение — единица
    # последнего знака, то есть тысячные. Отсюда две границы: у малых величин
    # шум сложения абсолютный, у больших он растёт вместе с самим числом, а
    # относительная граница остаётся на порядки ниже тысячных.
    SUM_MATCH_TOLERANCE = 1e-6
    SUM_MATCH_RELATIVE = 1e-13

    # Со скольких слагаемых запись диапазоном становится понятнее перечисления.
    # «=D7+E7» читается как подпись графы бланка «(гр.4+гр.5)», «=SUM(D7:E7)» —
    # уже нет; а вот сумма двадцати авиакомпаний перечислением нечитаема.
    SUM_RANGE_FROM = 4

    @staticmethod
    def _write_cell(ws, row: int, column: int, val: Any):
        """Пишет значение, сохраняя его тип: число — числом, текст — текстом."""
        if isinstance(val, Decimal):
            val = float(val)

        if isinstance(val, bool) or val is None:
            # bool — подкласс int, но в отчётности его нет; None — пустая ячейка.
            cell = ws.cell(row=row, column=column, value="" if val is None else str(val))
            cell.data_type = "s"
            return cell

        if isinstance(val, (int, float)):
            cell = ws.cell(row=row, column=column, value=val)
            cell.number_format = (
                ExportController.INT_FORMAT if float(val).is_integer()
                else ExportController.DECIMAL_FORMAT
            )
            cell.alignment = Alignment(horizontal="right", vertical="center")
            return cell

        text = str(val)
        cell = ws.cell(row=row, column=column, value=text)
        if text.startswith(ExportController.FORMULA_STARTERS):
            # openpyxl определяет формулу по ведущему символу строки. В ячейки
            # попадают названия из присланных файлов, то есть текст внешнего
            # происхождения: «=1+1» стало бы вычисляемой формулой, а «=cmd|…» —
            # приёмом внедрения формул на машине получателя (FUNC-9).
            cell.data_type = "s"
        return cell

    @staticmethod
    def _consecutive(numbers: list[int]) -> bool:
        """Идут ли номера подряд и по возрастанию — тогда это диапазон."""
        return all(b - a == 1 for a, b in zip(numbers, numbers[1:], strict=False))

    @staticmethod
    def _formula_text(operands: tuple[CellRef, ...], data_start_row: int) -> str:
        """Формула суммы по координатам слагаемых: «=D7+E7» или «=SUM(D7:M7)»."""
        refs = [
            f"{get_column_letter(col + 1)}{data_start_row + row}"
            for row, col in operands
        ]

        if len(operands) >= ExportController.SUM_RANGE_FROM:
            rows = [row for row, _ in operands]
            cols = [col for _, col in operands]
            one_line = (
                len(set(rows)) == 1 and ExportController._consecutive(cols)
                or len(set(cols)) == 1 and ExportController._consecutive(rows)
            )
            if one_line:
                return f"=SUM({refs[0]}:{refs[-1]})"

        return "=" + "+".join(refs)

    @staticmethod
    def _as_number(val: Any) -> float | None:
        """Значение ячейки числом; `None` — если складывать его нельзя."""
        if isinstance(val, Decimal):
            val = float(val)
        if isinstance(val, bool) or not isinstance(val, (int, float)):
            # bool в отчётности нет, а «Х» и «—» — это пометки бланка и отбора.
            return None
        return float(val)

    @staticmethod
    def _sum_checks_out(
        value: float, operands: tuple[CellRef, ...], rows: list[list[Any]]
    ) -> bool:
        """Складываются ли слагаемые ровно в это число.

        Сверка — не перестраховка от своих же ошибок. Итоги, которые свод считает
        сам, сходятся всегда; а строки 03, 07, 08 бланка 15-ГА, его графы «Всего»
        и тоннокилометраж 12-ГА приложение не считает, а хранит присланными.
        Формула там не показывает, откуда взялось число, а пересчитывает его
        заново — и если отчёт не сходится, в книге оказалась бы третья цифра,
        не совпадающая ни с экраном, ни с базой.

        Слагаемое, которого нет на листе или которое не число (пустая ячейка,
        «Х» неприменимой графы, «—» невыбранной), отменяет способ по той же
        причине: Excel считал бы такую ячейку нулём.
        """
        parts = []
        for operand_row, operand_col in operands:
            if not 0 <= operand_row < len(rows):
                return False
            operand_values = rows[operand_row]
            if not 0 <= operand_col < len(operand_values):
                return False
            part = ExportController._as_number(operand_values[operand_col])
            if part is None:
                return False
            parts.append(part)

        if not parts:
            return False

        # fsum, а не sum: слагаемых бывает две дюжины, и накопленная ошибка
        # обычного сложения сама по себе выглядела бы расхождением отчёта.
        return math.isclose(
            value,
            math.fsum(parts),
            rel_tol=ExportController.SUM_MATCH_RELATIVE,
            abs_tol=ExportController.SUM_MATCH_TOLERANCE,
        )

    @staticmethod
    def _apply_formula(
        cell,
        ways: tuple[tuple[CellRef, ...], ...],
        rows: list[list[Any]],
        row: int,
        col: int,
        data_start_row: int,
    ) -> bool:
        """Заменяет число формулой — первым способом, который даёт это же число.

        Способов у ячейки бывает два: сложить по строке и сложить по колонке.
        Не сошёлся ни один — остаётся число, как и было.
        """
        value = ExportController._as_number(rows[row][col])
        if value is None:
            return False

        for operands in ways:
            if ExportController._sum_checks_out(value, operands, rows):
                cell.value = ExportController._formula_text(operands, data_start_row)
                return True

        return False

    @staticmethod
    def _write_header(ws, header: ExportHeader | None) -> int:
        """Пишет шапку отчёта и возвращает номер строки, с которой идёт таблица.

        Без шапки книга получалась обезличенной: ни предприятия, ни периода, а
        лист назывался «Данные» (FUNC-4). Между шапкой и таблицей остаётся пустая
        строка — она же граница для автофильтра и для взгляда.
        """
        if header is None or not header.lines:
            return 1

        ws.title = header.sheet_title
        label_font = Font(bold=True)
        for offset, (label, value) in enumerate(header.lines):
            ws.cell(row=offset + 1, column=1, value=f"{label}:").font = label_font
            # Значение пишется текстом сознательно: это подпись отчёта, а не
            # величина, и превращать «12-ГА» в дату или число Excel не должен.
            cell = ws.cell(row=offset + 1, column=2, value=str(value))
            cell.data_type = "s"

        return len(header.lines) + 2

    @staticmethod
    def write_workbook(
        file_path: str,
        headers: list[str],
        rows: list[list[Any]],
        header_groups: list[tuple[int, int, str]] | None = None,
        header: ExportHeader | None = None,
        formulas: FormulaMap | None = None,
    ) -> None:
        """Пишет книгу. Об ошибке сообщает исключением, а не окном и не `False`.

        Прежде метод возвращал `False` и сам показывал `QMessageBox`: вызывающий
        не мог узнать, что именно не получилось, а тест не мог позвать экспорт
        вовсе — модальное окно остановило бы прогон (ARCH-2).

        `formulas` — ячейки, которые свод получил сложением, и способы их
        сложить. Каждый проходит сверку в `_apply_formula`: не сошлось ни одним —
        остаётся число, как и было.
        """
        ncols = len(headers)
        groups = header_groups or []
        cell_formulas = formulas or {}

        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hdr_font = Font(bold=True)

        # Заголовки таблицы начинаются под шапкой, а не с первой строки.
        top = ExportController._write_header(ws, header)
        header_rows = (top, top + 1) if groups else (top,)
        data_start_row: int

        if groups:
            in_group = set()
            for first, last, _ in groups:
                for c in range(first, last + 1):
                    in_group.add(c)

            # Верхний ряд: объединённые подписи месяцев / групп
            for first, last, label in groups:
                if first > last:
                    continue
                c1, c2 = first + 1, last + 1
                if c1 == c2:
                    cell = ws.cell(row=top, column=c1, value=label or "")
                else:
                    ws.merge_cells(
                        start_row=top, start_column=c1, end_row=top, end_column=c2
                    )
                    cell = ws.cell(row=top, column=c1, value=label or "")
                cell.alignment = center
                cell.font = hdr_font

            # Колонки без группы (напр. Показатель, Ед. изм., Код): заголовок на двух рядах
            for c in range(ncols):
                if c in in_group:
                    continue
                c1 = c + 1
                ws.merge_cells(start_row=top, start_column=c1, end_row=top + 1, end_column=c1)
                cell = ws.cell(row=top, column=c1, value=headers[c])
                cell.alignment = center
                cell.font = hdr_font

            # Нижний ряд заголовков для колонок внутри групп
            for c in range(ncols):
                if c not in in_group:
                    continue
                cell = ws.cell(row=top + 1, column=c + 1, value=headers[c])
                cell.alignment = center
                cell.font = hdr_font

            data_start_row = top + 2
        else:
            for c in range(ncols):
                cell = ws.cell(row=top, column=c + 1, value=headers[c])
                cell.alignment = center
                cell.font = hdr_font
            data_start_row = top + 1

        for r, row_values in enumerate(rows):
            excel_row = data_start_row + r
            for c in range(ncols):
                cell = ExportController._write_cell(ws, excel_row, c + 1, row_values[c])
                ways = cell_formulas.get((r, c))
                if ways:
                    ExportController._apply_formula(
                        cell, ways, rows, r, c, data_start_row
                    )

        # Ширина колонок — грубая оценка по тексту заголовка и первым строкам
        for c in range(ncols):
            letter = get_column_letter(c + 1)
            max_len = 10
            for check_row in header_rows:
                v = ws.cell(row=check_row, column=c + 1).value
                if v is not None:
                    max_len = max(max_len, min(60, len(str(v))))
            # Меряются значения, а не то, что записано в ячейку: в ячейке с
            # формулой лежит её текст, и «=SUM(D7:AB7)» растянуло бы колонку
            # тем длиннее, чем больше в сводке предприятий.
            for row_values in rows[:50]:
                v = row_values[c] if c < len(row_values) else None
                if v is not None:
                    max_len = max(max_len, min(60, len(str(v))))
            ws.column_dimensions[letter].width = max_len + 2

        wb.save(file_path)

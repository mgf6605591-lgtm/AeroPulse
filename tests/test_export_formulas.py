"""Формулы в выгруженной книге: итог остаётся суммой, а не отпечатком суммы.

В XLSX уходили только числа, и сложение, которым свод их получил, в файле не
оставалось. «Итого» ничем не отличался от набранной руками цифры: откуда он
взялся, по книге было не видно, а поправить в Excel одну колонку и получить
новый итог — нельзя.

Проверяются обе половины: правила построителей (что чем сложено) и книга (что
из этих правил дошло до файла). Отдельно — сверка: строки 03, 07, 08 бланка
15-ГА, его графы «Всего» и тоннокилометраж 12-ГА приложение не считает, а хранит
присланными, и формула там пересчитывала бы число заново. Не сходится отчёт —
в книге остаётся то же число, что на экране и в базе.

Книга собирается настоящим экспортёром и перечитывается `openpyxl`: проверяется
то, что окажется в файле у получателя.
"""

import os
import tempfile
import unittest
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker

from controllers.export_controller import ExportController
from controllers.report_filters import ReportFilters
from controllers.reports import ga15_airport, ga15_summary
from controllers.reports.formulas import PivotFormulas
from db.models.entities import Airport, Indicator, Locality
from tests.support import FakeRecord, MigratedDbCase, PivotCase
from utils.ga15_airport_layout import GA15_KEYS, GA15_METRIC_TAGS
from utils.ga15_summary_layout import GA15_SUMMARY_TOTAL_TITLE

try:
    from PyQt6.QtWidgets import QApplication, QTableView
    HAS_QT = True
except ImportError:  # PyQt6 отсутствует — проверки Qt пропускаются
    HAS_QT = False

_app = None


def setUpModule():
    global _app
    if HAS_QT:
        _app = QApplication.instance() or QApplication([])


# --------------------------------------------------------------------------
# Правила построителей: что свод объявляет суммой
# --------------------------------------------------------------------------


class Ga12FormulaRulesTest(PivotCase):
    """Своды 12-ГА объявляют суммой ровно то, что сами и сложили."""

    def records(self):
        return [
            FakeRecord("965", "Самолето-километры", "January", 2025, 100,
                       route_type="trunk"),
            FakeRecord("965", "Самолето-километры", "January", 2025, 50,
                       route_type="local"),
            FakeRecord("965", "Самолето-километры", "January", 2025, 20,
                       route_type="interregional"),
            FakeRecord("965", "Самолето-километры", "February", 2025, 7,
                       route_type="trunk"),
        ]

    def test_total_column_leaves_out_nested_route_types(self):
        """«Из них местные» входят во «Внутренние», и в сумме им не место."""
        result = self.build_per_airline_by_routes(self.records())

        self.assertEqual(
            ("m_2025_January_rt_trunk", "m_2025_January_rt_local"),
            result["formulas"].column_sums["m_2025_January_total"],
        )

    def test_total_column_rule_holds_for_every_period(self):
        result = self.build_per_airline_by_routes(self.records())

        self.assertIn("m_2025_February_total", result["formulas"].column_sums)

    def test_summary_column_sums_the_airlines_of_that_month(self):
        result = self.build_all_airlines(self.records())

        self.assertEqual(
            ("m_2025_January_a_0",),
            result["formulas"].column_sums["m_2025_January_total"],
        )

    def test_grand_total_sums_the_monthly_summaries(self):
        """Крайняя правая колонка складывает «Своды», а не авиакомпании заново."""
        result = self.build_all_airlines(self.records())

        self.assertEqual(
            ("m_2025_January_total", "m_2025_February_total"),
            result["formulas"].column_sums["grand_total"],
        )

    def test_airline_period_total_sums_its_own_months(self):
        """Итог предприятия за период складывает его же колонки, а не «Своды»."""
        result = self.build_all_airlines(self.records())

        self.assertEqual(
            ("m_2025_January_a_0", "m_2025_February_a_0"),
            result["formulas"].column_sums["total_a_0"],
        )

    def test_collapsed_summary_sums_the_periods(self):
        result = self.build_per_airline_summary(self.records())

        self.assertEqual(
            ("m_2025_January", "m_2025_February"),
            result["formulas"].column_sums["total"],
        )

    def test_multi_airline_totals_stay_within_their_airline(self):
        result = self.build_multi_airline_by_routes(self.records())

        rule = result["formulas"].column_sums["m_2025_January_aid_1_total"]

        self.assertTrue(all("aid_1_" in key for key in rule), rule)


class Ga12TonneKilometreRuleTest(PivotCase):
    """Тоннокилометраж: строка бланка и три её строки «в том числе»."""

    def setUp(self):
        super().setUp()
        with self.Session() as session:
            for code, name in (
                ("450", "Выполненный тоннокилометраж"),
                ("450пас", "а) пассажирский"),
                ("450гр", "б) грузовой"),
                ("450пч", "в) почтовый"),
            ):
                session.add(Indicator(name=name, code=code, measure="тыс. ткм"))
            session.commit()

    def result(self):
        return self.build_per_airline_summary([
            FakeRecord("450", "Выполненный тоннокилометраж", "January", 2025, 90,
                       measure="тыс. ткм"),
            FakeRecord("450пас", "а) пассажирский", "January", 2025, 60,
                       measure="тыс. ткм"),
            FakeRecord("450гр", "б) грузовой", "January", 2025, 20,
                       measure="тыс. ткм"),
            FakeRecord("450пч", "в) почтовый", "January", 2025, 10,
                       measure="тыс. ткм"),
        ])

    def index_of(self, result, code):
        for index, row in enumerate(result["rows"]):
            if row.get("code") == code:
                return index
        raise AssertionError(f"в своде нет строки с кодом {code}")

    def test_parent_row_sums_its_three_detail_rows(self):
        result = self.result()

        self.assertEqual(
            tuple(self.index_of(result, code) for code in ("450пас", "450гр", "450пч")),
            result["formulas"].row_sums[self.index_of(result, "450")],
        )

    def test_detail_rows_themselves_sum_nothing(self):
        """Слагаемое — не итог: у строк «в том числе» своего правила нет."""
        result = self.result()

        self.assertNotIn(self.index_of(result, "450пас"), result["formulas"].row_sums)


class FakeAirportAggregateRow:
    """Строка агрегата 15-ГА — то, что отдаёт база после GROUP BY."""

    def __init__(self, indicator_code: str, total, airport_id: int = 1,
                 month: str = "January", year: int = 2025, records: int = 1):
        self.indicator_code = indicator_code
        self.total = total
        self.airport_id = airport_id
        self.month = month
        self.year = year
        self.records = records


GA15_ROW_CODES = ("R01", "R02", "R03", "R04ИНО", "R05", "R06", "R07", "R08", "R09")


class Ga15BlankRulesTest(MigratedDbCase):
    """Бланк 15-ГА: графы «Всего» и строки 03, 07, 08."""

    def setUp(self):
        super().setUp()
        Session = sessionmaker(bind=self.engine, expire_on_commit=False)
        with Session() as session:
            session.add(Locality(id=1, name="Город"))
            session.add(Airport(id=1, code="ЯКТ", name="Якутск", locality_id=1))
            session.commit()
        patcher = patch("controllers.reports.ga15_airport.get_session", Session)
        patcher.start()
        self.addCleanup(patcher.stop)

    def build(self, rows=()):
        with patch("controllers.airport_ind_service.AirportIndicatorService.aggregate",
                   return_value=list(rows)):
            return ga15_airport.build(ReportFilters(), airport_id=1)

    @staticmethod
    def key(tag: str) -> str:
        return GA15_KEYS[2 + GA15_METRIC_TAGS.index(tag)]

    def line_index(self, result, line: str) -> int:
        for index, row in enumerate(result["rows"]):
            if str(row.get(GA15_KEYS[1])) == line:
                return index
        raise AssertionError(f"в бланке нет строки {line}")

    def test_passenger_total_graph_sums_sent_and_received(self):
        rule = self.build()["formulas"].column_sums[self.key("ПАС_ВСЕГО")]

        self.assertEqual((self.key("ПАС_ОТП"), self.key("ПАС_ПРИН")), rule)

    def test_every_total_graph_of_the_blank_has_a_rule(self):
        column_sums = self.build()["formulas"].column_sums

        self.assertEqual(
            [self.key(tag) for tag in ("ПАС_ВСЕГО", "ГР_ВСЕГО", "ПЧ_ВСЕГО")],
            sorted(column_sums, key=GA15_KEYS.index),
        )

    def test_commercial_total_row_sums_lines_03_and_07(self):
        """Строка 08 подписана «(стр. 03 + стр. 07)» — так и записано правило."""
        result = self.build()

        self.assertEqual(
            (self.line_index(result, "3"), self.line_index(result, "7")),
            result["formulas"].row_sums[self.line_index(result, "8")],
        )

    def test_the_line_number_is_not_a_sum_of_line_numbers(self):
        """Номер строки — имя строки, а не величина, и складывать его нечего.

        Правило строки касается всех колонок сразу, и сверка сама по себе тут не
        спасает: строка 03 подписана «(стр. 01+стр. 02)», и 1 + 2 сошлось ровно в
        её собственный номер — в выгрузке он становился вычисляемым.
        """
        result = self.build()
        rule = result["formulas"].operands(
            self.line_index(result, "3"),
            1,
            GA15_KEYS[1],
            {key: index for index, key in enumerate(GA15_KEYS)},
        )

        self.assertEqual((), rule)

    def test_line_04_stays_out_of_the_international_total(self):
        """«В том числе иностранными» — из строки 03, а не рядом с ней."""
        result = self.build()

        self.assertEqual(
            (self.line_index(result, "1"), self.line_index(result, "2")),
            result["formulas"].row_sums[self.line_index(result, "3")],
        )


class LabelColumnTest(unittest.TestCase):
    """Колонка-подпись не складывается, даже когда сложение сошлось."""

    RULE = PivotFormulas(
        row_sums={2: (0, 1)}, label_keys=frozenset({"number"})
    )
    COLUMNS = {"number": 0, "value": 1}

    def test_a_label_column_has_no_way_to_be_summed(self):
        self.assertEqual(
            (), self.RULE.operands(2, 0, "number", self.COLUMNS)
        )

    def test_the_value_column_beside_it_still_has_one(self):
        self.assertEqual(
            (((0, 1), (1, 1)),), self.RULE.operands(2, 1, "value", self.COLUMNS)
        )


class Ga15SummaryRulesTest(MigratedDbCase):
    """Сводка по аэропортам: кварталы, нарастающие итоги и строка «Итого»."""

    ENTERPRISE, CHILD = 1, 2

    def setUp(self):
        super().setUp()
        Session = sessionmaker(bind=self.engine, expire_on_commit=False)
        with Session() as session:
            session.add(Locality(id=1, name="Город"))
            session.add(Airport(id=self.ENTERPRISE, code="FKP", name="ФКП",
                                locality_id=1))
            session.add(Airport(id=self.CHILD, code="ALDAN", name="Алдан",
                                locality_id=1, parent_id=self.ENTERPRISE))
            session.commit()
        patcher = patch("controllers.reports.ga15_summary.get_session", Session)
        patcher.start()
        self.addCleanup(patcher.stop)

    def build(self, **filters):
        base = {"period_from": (2025, 1), "period_to": (2025, 3)}
        base.update(filters)
        with patch("controllers.airport_ind_service.AirportIndicatorService.aggregate",
                   return_value=[]):
            return ga15_summary.build(ReportFilters(**base))

    def test_quarter_column_sums_its_three_months(self):
        rule = self.build()["formulas"].column_sums["q2025_1_ВС"]

        self.assertEqual(("m2025_1_ВС", "m2025_2_ВС", "m2025_3_ВС"), rule)

    def test_month_total_graph_sums_sent_and_received(self):
        rule = self.build()["formulas"].column_sums["m2025_1_ПАС_ВСЕГО"]

        self.assertEqual(("m2025_1_ПАС_ОТП", "m2025_1_ПАС_ПРИН"), rule)

    def test_quarter_total_graph_prefers_its_own_months(self):
        """У квартала сумма месяцев вернее, чем сумма двух других сумм."""
        rule = self.build()["formulas"].column_sums["q2025_1_ПАС_ВСЕГО"]

        self.assertEqual(
            ("m2025_1_ПАС_ВСЕГО", "m2025_2_ПАС_ВСЕГО", "m2025_3_ПАС_ВСЕГО"), rule
        )

    def test_total_row_sums_enterprises_without_their_breakdown(self):
        """Разбивка предприятия уже вошла в итог его собственной строкой."""
        result = self.build()
        titles = [row["airport"].strip() for row in result["rows"]]
        total_index = titles.index(GA15_SUMMARY_TOTAL_TITLE)

        self.assertEqual((titles.index("ФКП"),), result["formulas"].row_sums[total_index])

    def test_summary_without_airports_promises_no_sums(self):
        """Отобран аэропорт, которого нет: складывать в такой сводке нечего."""
        result = self.build(airport_ids=(999,))

        self.assertFalse(result["formulas"])


# --------------------------------------------------------------------------
# Книга: что из правил дошло до файла
# --------------------------------------------------------------------------

HEADERS = ["Показатель", "Международные", "Внутренние", "из них местные", "ИТОГО"]
KEYS = ["indicator", "trunk", "local", "inner", "total"]


@unittest.skipUnless(HAS_QT, "PyQt6 не установлен")
class WorkbookCase(unittest.TestCase):
    """Книга собирается настоящим экспортом с настоящей моделью свода."""

    def setUp(self):
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        self.path = os.path.join(tmp.name, "export.xlsx")

    def export(self, rows, formulas, headers=None, keys=None, header_groups=None):
        from forms.models.pivot_dict_model import PivotDictModel
        from forms.table_export import export_table_to_excel

        model = PivotDictModel()
        model.set_source_data(rows, headers or HEADERS, keys or KEYS, formulas)
        view = QTableView()
        self.addCleanup(view.deleteLater)
        view.setModel(model)

        self.assertTrue(
            export_table_to_excel(view, self.path, header_groups=header_groups)
        )
        return load_workbook(self.path).active


class CheckedSumTest(WorkbookCase):
    """Формула пишется там, где даёт ровно то же число."""

    RULE = PivotFormulas(column_sums={"total": ("trunk", "local")})

    def row(self, total, trunk=100.0, local=50.0, inner=20.0):
        return [{"indicator": "Самолето-километры", "trunk": trunk, "local": local,
                 "inner": inner, "total": total}]

    def test_matching_sum_becomes_a_formula(self):
        ws = self.export(self.row(150.0), self.RULE)

        self.assertEqual("f", ws["E2"].data_type)
        self.assertEqual("=B2+C2", ws["E2"].value)

    def test_mismatching_sum_stays_the_number_from_the_screen(self):
        """Бланк не сошёлся — в книге то же число, что на экране, а не третье."""
        ws = self.export(self.row(151.0), self.RULE)

        self.assertEqual("n", ws["E2"].data_type)
        self.assertEqual(151, ws["E2"].value)

    def test_a_mark_instead_of_an_addend_cancels_the_formula(self):
        """«—» невыбранной графы Excel считал бы нулём."""
        ws = self.export(self.row(100.0, local="—"), self.RULE)

        self.assertEqual("n", ws["E2"].data_type)

    def test_empty_addend_cancels_the_formula(self):
        ws = self.export(self.row(100.0, local=None), self.RULE)

        self.assertEqual("n", ws["E2"].data_type)

    def test_rounding_noise_still_counts_as_matching(self):
        """0.1 + 0.2 в двоичном сложении — не 0.3, и расхождением это не считается."""
        ws = self.export(self.row(0.3, trunk=0.1, local=0.2), self.RULE)

        self.assertEqual("f", ws["E2"].data_type)

    def test_a_hundredth_apart_is_a_discrepancy_and_not_noise(self):
        ws = self.export(self.row(150.01), self.RULE)

        self.assertEqual("n", ws["E2"].data_type)

    def test_formula_cell_keeps_the_number_format(self):
        ws = self.export(self.row(150.0), self.RULE)

        self.assertEqual(ExportController.INT_FORMAT, ws["E2"].number_format)
        self.assertEqual("right", ws["E2"].alignment.horizontal)


class FormulaShapeTest(WorkbookCase):
    """Как формула записана: перечислением или диапазоном."""

    def wide_export(self, count):
        headers = ["Показатель"] + [f"АК {i}" for i in range(count)] + ["Свод"]
        keys = ["indicator"] + [f"a_{i}" for i in range(count)] + ["total"]
        row = {"indicator": "Самолето-километры", "total": float(count)}
        row.update({f"a_{i}": 1.0 for i in range(count)})
        rule = PivotFormulas(
            column_sums={"total": tuple(f"a_{i}" for i in range(count))}
        )
        return self.export([row], rule, headers=headers, keys=keys)

    def test_two_addends_are_spelled_out(self):
        """«=B2+C2» читается как подпись графы бланка «(гр.4+гр.5)»."""
        ws = self.wide_export(2)

        self.assertEqual("=B2+C2", ws["D2"].value)

    def test_a_long_run_of_neighbours_becomes_a_range(self):
        ws = self.wide_export(6)

        self.assertEqual("=SUM(B2:G2)", ws["H2"].value)

    def test_gaps_keep_the_formula_spelled_out(self):
        """Пропуск в ряду диапазоном не запишешь: в него попало бы лишнее."""
        rows = [{"indicator": "Самолето-километры", "trunk": 100.0, "local": 50.0,
                 "inner": 20.0, "total": 150.0}]
        ws = self.export(rows, PivotFormulas(column_sums={"total": ("trunk", "local")}))

        self.assertEqual("=B2+C2", ws["E2"].value)

    def test_a_column_of_rows_becomes_a_range_too(self):
        rows = [
            {"indicator": f"Аэропорт {i}", "trunk": 1.0, "local": 0.0, "inner": 0.0,
             "total": 1.0}
            for i in range(4)
        ]
        rows.append({"indicator": "Итого:", "trunk": 4.0, "local": 0.0, "inner": 0.0,
                     "total": 4.0})
        ws = self.export(rows, PivotFormulas(row_sums={4: (0, 1, 2, 3)}))

        self.assertEqual("=SUM(B2:B5)", ws["B6"].value)


class FallbackWayTest(WorkbookCase):
    """У ячейки бывает два способа сложения — годится первый сошедшийся."""

    def rows(self):
        return [
            {"indicator": "Якутск", "trunk": 100.0, "local": 50.0, "inner": 0.0,
             "total": 151.0},
            {"indicator": "Итого:", "trunk": 100.0, "local": 50.0, "inner": 0.0,
             "total": 151.0},
        ]

    RULE = PivotFormulas(
        column_sums={"total": ("trunk", "local")}, row_sums={1: (0,)}
    )

    def test_the_row_rule_saves_the_cell_when_the_column_rule_falls_apart(self):
        """Графа бланка не сошлась, но итог по строке от этого не перестал быть итогом."""
        ws = self.export(self.rows(), self.RULE)

        self.assertEqual("=E2", ws["E3"].value)

    def test_the_column_rule_goes_first_where_both_check_out(self):
        rows = self.rows()
        rows[0]["total"] = rows[1]["total"] = 150.0
        ws = self.export(rows, self.RULE)

        self.assertEqual("=B3+C3", ws["E3"].value)


class SheetOffsetTest(WorkbookCase):
    """Ссылки считаются от таблицы, а не от первой строки листа."""

    def test_references_account_for_the_two_header_rows(self):
        rows = [{"indicator": "Самолето-километры", "trunk": 100.0, "local": 50.0,
                 "inner": 20.0, "total": 150.0}]
        ws = self.export(
            rows,
            PivotFormulas(column_sums={"total": ("trunk", "local")}),
            header_groups=[(1, 4, "Январь 2025")],
        )

        self.assertEqual("=B3+C3", ws["E3"].value)

    def test_column_width_ignores_the_length_of_the_formula(self):
        """В ячейке лежит текст формулы, и мерить колонку по нему нельзя.

        Слагаемые здесь идут через одно, поэтому формула записывается
        перечислением и оказывается вдвое длиннее самого числа.
        """
        headers = ["Показатель"] + [f"АК {i}" for i in range(12)] + ["Свод"]
        keys = ["indicator"] + [f"a_{i}" for i in range(12)] + ["total"]
        row = {"indicator": "Самолето-километры", "total": 6.0}
        row.update({f"a_{i}": float(i % 2 == 0) for i in range(12)})
        rule = PivotFormulas(
            column_sums={"total": tuple(f"a_{i}" for i in range(0, 12, 2))}
        )
        ws = self.export([row], rule, headers=headers, keys=keys)

        self.assertEqual(18, len(ws["N2"].value))
        self.assertLessEqual(ws.column_dimensions["N"].width, len("Свод") + 12)


class ModelWithoutRulesTest(WorkbookCase):
    """Свод без правил выгружается ровно так же, как выгружался раньше."""

    def test_numbers_stay_numbers_when_nothing_is_declared_a_sum(self):
        rows = [{"indicator": "Самолето-километры", "trunk": 100.0, "local": 50.0,
                 "inner": 20.0, "total": 150.0}]
        ws = self.export(rows, None)

        self.assertEqual("n", ws["E2"].data_type)
        self.assertEqual(150, ws["E2"].value)


class WithoutQtTest(unittest.TestCase):
    """Книга с формулами собирается и без Qt: правила приходят готовыми (ARCH-2)."""

    def setUp(self):
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        self.path = os.path.join(tmp.name, "export.xlsx")

    def write(self, formulas):
        ExportController.write_workbook(
            self.path,
            headers=["Показатель", "Январь", "Февраль", "Всего"],
            rows=[["Самолето-километры", 100.0, 50.0, 150.0]],
            formulas=formulas,
        )
        return load_workbook(self.path).active

    def test_formula_reaches_the_file(self):
        ws = self.write({(0, 3): (((0, 1), (0, 2)),)})

        self.assertEqual("=B2+C2", ws["D2"].value)

    def test_reference_outside_the_table_is_refused(self):
        """Ссылка в никуда дала бы в Excel ноль вместо числа."""
        ws = self.write({(0, 3): (((0, 1), (7, 2)),)})

        self.assertEqual(150, ws["D2"].value)


if __name__ == "__main__":
    unittest.main()

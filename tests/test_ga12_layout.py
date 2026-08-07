"""Раскладка бланка 12-ГА: строка формы, её код и значения по графам.

Проверяется класс дефектов, при котором импорт проходит без единой жалобы, а в
базу ложатся правдоподобные, но чужие цифры:

* строки листа адресовались жёсткими индексами, и на настоящем бланке они были
  смещены на единицу — каждое значение уходило под код соседнего показателя;
* детализация тоннокилометража не читалась из XLSX вовсе, зато из XLSX читались
  восемь строк некоммерческих полётов, которых в бланке нет, — один и тот же
  отчёт в двух форматах давал в базе разный набор строк (BUG-3).

Проверки на настоящем бланке лежат в `test_import_gate.py` и пропускаются, если
файла нет; здесь книги и XML собираются синтетически.
"""

import os
import tempfile
import unittest
import xml.etree.ElementTree as ET

from openpyxl import Workbook

from parsers.xlsx_parser import XLSXParser
from parsers.xml_parser import XMLParser
from tests.support import (
    GA12_ROWS as BLANK_FIXTURE_ROWS,
    ga12_cell_value,
    ga12_total_value,
    make_ga12_workbook,
)
from utils.constants import (
    GA12_CODE_ORDER_FLAT,
    GA12_CODES_BY_SECTION,
    GA12_DETAIL_TON_CODES,
)
from utils.ga12_layout import (
    GA12_ROW_BY_BLANK_NUMBER,
    GA12_ROW_BY_XML_ROW,
    GA12_ROWS,
    SECTION_NON_COMMERCIAL,
)

# Значение строки в графе: одно и то же число для обоих форматов, своё для каждой
# графы — перепутанные графы сразу видны.
def cell_value(row_index: int, col: int) -> float:
    return row_index * 100 + col


def make_full_ga12_workbook(path: str) -> str:
    """Книга со всеми строками бланка — по описанию формы из utils.ga12_layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ГА12"
    ws.cell(row=1, column=1, value="Форма 12-ГА")

    excel_row = 11
    for index, row in enumerate(GA12_ROWS):
        if row.detail_marker:
            ws.cell(row=excel_row, column=1, value=f"            {row.detail_marker} {row.name}")
        else:
            ws.cell(row=excel_row, column=1, value=row.name)
            ws.cell(row=excel_row, column=2, value=row.blank_number)
        ws.cell(row=excel_row, column=4, value=row.okei)
        for col in range(5, 10):  # столбцы листа: графа бланка на единицу меньше
            ws.cell(row=excel_row, column=col, value=cell_value(index, col - 1))
        excel_row += 1

    wb.save(path)
    return path


def make_full_ga12_xml(path: str) -> str:
    """XML-выгрузка того же отчёта: те же строки бланка с теми же значениями."""
    root = ET.Element("report", {"year": "2025", "period": "01"})
    title = ET.SubElement(root, "title")
    ET.SubElement(title, "item", {"name": "name", "value": "Тестовая АК"})
    section = ET.SubElement(ET.SubElement(root, "sections"), "section")

    for index, row in enumerate(GA12_ROWS):
        node = ET.SubElement(section, "row", {"code": str(row.xml_row)})
        for col in range(4, 9):  # графы 4…8 бланка
            ET.SubElement(node, "col", {"code": str(col)}).text = str(cell_value(index, col))

    ET.ElementTree(root).write(path, encoding="utf-8", xml_declaration=True)
    return path


def by_code(result: dict) -> dict:
    """Значения по (код показателя, вид сообщения)."""
    return {(i["indicator_code"], i["route_type"]): i["value"] for i in result["indicators"]}


class WorkbookCase(unittest.TestCase):
    def setUp(self):
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        self.tmp_dir = tmp.name

    def path(self, name: str) -> str:
        return os.path.join(self.tmp_dir, name)


class BlankRowMappingTest(WorkbookCase):
    """Значение строки бланка попадает под код именно этой строки."""

    def setUp(self):
        super().setUp()
        self.values = by_code(XLSXParser.parse_file(make_ga12_workbook(self.path("a.xlsx"))))

    def test_first_row_keeps_its_own_code(self):
        """Строка 1 бланка — «Самолето-километры», код 965.

        Прежняя раскладка начинала отсчёт со строки заголовка раздела, поэтому
        значения строки 1 попадали в базу под кодом 642 «Отправлений воздушных
        судов», значения строки 2 — под 356 «Налет часов» и так до конца бланка.
        """
        # Графы 4 и 5 хранятся суммой: международные — это два столбца бланка.
        self.assertEqual(ga12_cell_value(1, 5) + ga12_cell_value(1, 6), self.values[("965", "trunk")])
        self.assertEqual(ga12_cell_value(1, 7), self.values[("965", "local")])
        self.assertEqual(ga12_cell_value(1, 8), self.values[("965", "interregional")])
        self.assertEqual(ga12_cell_value(1, 9), self.values[("965", "subsidir")])

    def test_each_numbered_row_keeps_its_own_code(self):
        for number, code in ((2, "642"), (3, "356"), (4, "792"), (7, "423"), (9, "450")):
            with self.subTest(code=code):
                self.assertEqual(ga12_cell_value(number, 7), self.values[(code, "local")])

    def test_column_number_row_is_not_a_data_row(self):
        """Служебная строка «1|2|3|…|9» под шапкой не должна попасть в данные.

        В её графе 2 стоит число 2, по которому она неотличима от строки 2 бланка,
        а в графах данных — номера граф 4…8.
        """
        self.assertEqual(ga12_cell_value(2, 7), self.values[("642", "local")])

    def test_ton_detail_rows_are_read(self):
        """Строки «в том числе» номера не имеют и опознаются по маркеру а)/б)/в)."""
        for code in GA12_DETAIL_TON_CODES:
            with self.subTest(code=code):
                self.assertIn((code, "local"), self.values)

    def test_total_column_is_not_imported(self):
        """Графа 9 «ИТОГО гр.4+гр.5+гр.6» производная: в базе она удвоила бы отчёт."""
        totals = {ga12_total_value(number) for number, _, _ in BLANK_FIXTURE_ROWS}
        self.assertFalse(totals & set(self.values.values()))


class FormatParityTest(WorkbookCase):
    """Один отчёт в XLSX и в XML даёт в базе одно и то же (BUG-3)."""

    def setUp(self):
        super().setUp()
        self.from_xlsx = XLSXParser.parse_file(make_full_ga12_workbook(self.path("a.xlsx")))
        self.from_xml = XMLParser.parse_file(make_full_ga12_xml(self.path("a.xml")))

    def test_same_set_of_rows(self):
        self.assertEqual(
            {i["indicator_code"] for i in self.from_xml["indicators"]},
            {i["indicator_code"] for i in self.from_xlsx["indicators"]},
        )

    def test_same_values(self):
        self.assertEqual(by_code(self.from_xml), by_code(self.from_xlsx))

    def test_same_names_and_measures(self):
        def described(result):
            return {
                (i["indicator_code"], i["indicator_name"], i["measure"], i["regularity"])
                for i in result["indicators"]
            }

        self.assertEqual(described(self.from_xml), described(self.from_xlsx))

    def test_whole_blank_is_read(self):
        self.assertEqual(
            {row.code for row in GA12_ROWS},
            {i["indicator_code"] for i in self.from_xlsx["indicators"]},
        )


class OkeiCrossCheckTest(WorkbookCase):
    """Графа «Код по ОКЕИ» сверяется с бланком."""

    def test_foreign_okei_is_refused(self):
        """Чужой код означает другую раскладку строк — это отказ, а не пропуск строки.

        Пропустить молча нельзя: отчёт уйдёт в базу неполным и без признаков этого.
        """
        path = make_ga12_workbook(
            self.path("a.xlsx"),
            labels=((1, "Самолето-километры", "111"),) + tuple(
                row for row in [(2, "Отправлений воздушных судов", "642"),
                                (3, "Налет часов", "356"),
                                (4, "Перевезено пассажиров", "792"),
                                (7, "Выполненный пассажирооборот", "423")]
            ),
        )
        with self.assertRaises(ValueError) as ctx:
            XLSXParser.parse_file(path)
        self.assertIn("ОКЕИ", str(ctx.exception))

    def test_empty_okei_column_is_allowed(self):
        """Пустая графа — не повод отказывать: сверять нечего."""
        path = make_ga12_workbook(self.path("b.xlsx"))
        from openpyxl import load_workbook

        wb = load_workbook(path)
        for excel_row in range(11, 25):
            wb["ГА12"].cell(row=excel_row, column=4, value=None)
        wb.save(path)

        self.assertTrue(XLSXParser.parse_file(path)["indicators"])


class LayoutTableTest(unittest.TestCase):
    """Сама таблица бланка: ключи уникальны, разделы согласованы (ARCH-12)."""

    def test_codes_are_unique(self):
        codes = [row.code for row in GA12_ROWS]
        self.assertEqual(len(codes), len(set(codes)))

    def test_blank_numbers_are_unique(self):
        numbers = [row.blank_number for row in GA12_ROWS if row.blank_number is not None]
        self.assertEqual(len(numbers), len(set(numbers)))
        self.assertEqual(len(numbers), len(GA12_ROW_BY_BLANK_NUMBER))

    def test_xml_rows_are_unique(self):
        self.assertEqual(len(GA12_ROWS), len(GA12_ROW_BY_XML_ROW))

    def test_sections_cover_the_whole_form(self):
        """Разделы больше не нарезаются из плоского списка срезами по индексам."""
        from_sections = [code for codes in GA12_CODES_BY_SECTION.values() for code in codes]
        self.assertEqual(GA12_CODE_ORDER_FLAT, from_sections)

    def test_non_commercial_section_matches_the_blank(self):
        """В бланке и в метаформе XML это одна строка — «Налет часов».

        Прежняя раскладка XLSX объявляла девять строк некоммерческих полётов;
        восемь из них в форме отсутствуют, а под первую попадала единственная
        настоящая строка раздела.
        """
        codes = [row.code for row in GA12_ROWS if row.section == SECTION_NON_COMMERCIAL]
        self.assertEqual(["356нк"], codes)

    def test_detail_rows_point_at_an_existing_parent(self):
        codes = {row.code for row in GA12_ROWS}
        for row in GA12_ROWS:
            if row.detail_of:
                with self.subTest(code=row.code):
                    self.assertIn(row.detail_of, codes)


if __name__ == "__main__":
    unittest.main()

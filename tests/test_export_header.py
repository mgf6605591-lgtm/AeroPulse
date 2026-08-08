"""Шапка выгружаемой книги (FUNC-4).

В файл уходили только заголовки колонок и цифры, а лист назывался «Данные».
Выгруженный отчёт 12-ГА получался обезличенным: определить по нему авиакомпанию
и период было нельзя, и две выгрузки различались разве что именем файла, которое
пользователь задавал вручную. Вместе с DATA-1, где колонки подписаны месяцем без
года, файл переставал читаться вне того сеанса, в котором его сделали.

Сборка шапки проверяется напрямую — она не знает ни о Qt, ни о книге. Запись
проверяется настоящим экспортёром с перечитыванием файла `openpyxl`: сдвиг
таблицы под шапку — как раз то место, где легко разъехаться заголовкам.
"""

import os
import tempfile
import unittest
from datetime import datetime

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook

from controllers.export_controller import ExportController
from controllers.export_header import build_export_header, period_label
from utils.constants import MODE_AIRLINE, MODE_AIRPORT, VIEW_DETAIL, VIEW_PIVOT

try:
    from PyQt6.QtWidgets import QApplication, QTableView
    HAS_QT = True
except ImportError:  # PyQt6 отсутствует — проверки Qt пропускаются
    HAS_QT = False

_app = None


def setUpModule():
    global _app
    if HAS_QT:
        _app = QApplication.instance() or QApplication([])


MOMENT = datetime(2026, 8, 8, 14, 5)

HEADERS = ["Показатель", "Январь 2025", "Февраль 2025"]
KEYS = ["indicator", "m_2025_January", "m_2025_February"]
ROWS = [
    {"indicator": "Самолето-километры", "m_2025_January": 1234567.0, "m_2025_February": 1234.567},
    {"indicator": "Налет часов", "m_2025_January": 0.0, "m_2025_February": None},
]


def header_for(**kwargs):
    kwargs.setdefault("mode", MODE_AIRLINE)
    kwargs.setdefault("view", VIEW_PIVOT)
    kwargs.setdefault("now", MOMENT)
    return build_export_header(**kwargs)


def value_of(header, label):
    for name, value in header.lines:
        if name == label:
            return value
    return None


class HeaderTellsWhatTheReportIsTest(unittest.TestCase):
    """Всё, чего в файле не было: форма, предприятие, период, момент выгрузки."""

    def test_form_follows_the_tab(self):
        self.assertEqual("12-ГА", value_of(header_for(mode=MODE_AIRLINE), "Форма"))
        self.assertEqual("15-ГА", value_of(header_for(mode=MODE_AIRPORT), "Форма"))

    def test_enterprise_comes_from_the_report(self):
        header = header_for(stats={"airline_name": "Тестовая АК"})

        self.assertEqual("Тестовая АК", value_of(header, "Предприятие"))

    def test_airport_report_names_the_airport(self):
        header = header_for(mode=MODE_AIRPORT, stats={"airport_name": "Якутск"})

        self.assertEqual("Якутск", value_of(header, "Предприятие"))

    def test_summary_over_many_says_so(self):
        """У свода по всем предприятиям названия нет — и пустая строка тут хуже прочерка."""
        self.assertEqual("свод по всем авиакомпаниям",
                         value_of(header_for(stats={}), "Предприятие"))
        self.assertEqual("свод по всем аэропортам",
                         value_of(header_for(mode=MODE_AIRPORT, stats={}), "Предприятие"))

    def test_moment_of_export_is_written_down(self):
        self.assertEqual("08.08.2026 14:05", value_of(header_for(), "Выгружено"))

    def test_user_is_named_when_known(self):
        self.assertEqual("ваня", value_of(header_for(user="ваня"), "Пользователь"))

    def test_unknown_user_leaves_no_empty_line(self):
        self.assertIsNone(value_of(header_for(), "Пользователь"))

    def test_view_is_named(self):
        self.assertEqual("свод", value_of(header_for(view=VIEW_PIVOT), "Представление"))
        self.assertEqual("подробная таблица",
                         value_of(header_for(view=VIEW_DETAIL), "Представление"))


class PeriodLabelTest(unittest.TestCase):
    """Период — то, чего в файле не хватало сильнее всего: колонки подписаны
    месяцем без года (DATA-1), и вне контекста отчёт не читался."""

    def test_single_month_is_not_repeated_twice(self):
        label = period_label({"period_from": (2025, 1), "period_to": (2025, 1)})

        self.assertEqual("Январь 2025", label)

    def test_range_shows_both_ends(self):
        label = period_label({"period_from": (2025, 1), "period_to": (2025, 3)})

        self.assertEqual("Январь 2025 — Март 2025", label)

    def test_range_across_years(self):
        label = period_label({"period_from": (2024, 12), "period_to": (2025, 2)})

        self.assertEqual("Декабрь 2024 — Февраль 2025", label)

    def test_missing_period_gives_no_line(self):
        self.assertEqual("", period_label({}))
        self.assertIsNone(value_of(header_for(filters={}), "Период"))

    def test_period_reaches_the_header(self):
        header = header_for(filters={"period_from": (2025, 1), "period_to": (2025, 3)})

        self.assertEqual("Январь 2025 — Март 2025", value_of(header, "Период"))


class CountersMatchTheScreenTest(unittest.TestCase):
    def test_counters_are_carried_over(self):
        header = header_for(stats={"indicators": 31, "records": 210})

        self.assertEqual("31", value_of(header, "Показателей"))
        self.assertEqual("210", value_of(header, "Записей"))

    def test_absent_counters_leave_no_empty_lines(self):
        header = header_for(stats={})

        self.assertIsNone(value_of(header, "Показателей"))
        self.assertIsNone(value_of(header, "Записей"))


class SheetTitleTest(unittest.TestCase):
    def test_sheet_is_named_after_the_form(self):
        self.assertEqual("12-ГА (свод)", header_for().sheet_title)

    def test_detail_view_has_its_own_name(self):
        self.assertEqual("15-ГА (подробная таблица)",
                         header_for(mode=MODE_AIRPORT, view=VIEW_DETAIL).sheet_title)

    def test_title_fits_the_excel_limit(self):
        """Excel обрезает название листа длиннее 31 символа и ругается на файл."""
        for mode in (MODE_AIRLINE, MODE_AIRPORT):
            for view in (VIEW_PIVOT, VIEW_DETAIL):
                title = header_for(mode=mode, view=view).sheet_title
                self.assertLessEqual(len(title), 31, title)


@unittest.skipUnless(HAS_QT, "PyQt6 не установлен")
class WorkbookCase(unittest.TestCase):
    def setUp(self):
        from forms.models.pivot_dict_model import PivotDictModel

        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        self.path = os.path.join(tmp.name, "export.xlsx")

        self.model = PivotDictModel()
        self.model.set_source_data(ROWS, HEADERS, KEYS)
        self.view = QTableView()
        self.addCleanup(self.view.deleteLater)
        self.view.setModel(self.model)

    def export(self, header=None, header_groups=None):
        self.assertTrue(ExportController.export_to_excel(
            self.view, self.path, header_groups=header_groups, header=header
        ))
        return load_workbook(self.path).active


@unittest.skipUnless(HAS_QT, "PyQt6 не установлен")
class HeaderReachesTheFileTest(WorkbookCase):
    """То, ради чего пункт и заведён: по файлу видно, что это за отчёт."""

    def setUp(self):
        super().setUp()
        self.header = header_for(
            stats={"airline_name": "Тестовая АК", "indicators": 31},
            filters={"period_from": (2025, 1), "period_to": (2025, 3)},
            user="ваня",
        )

    def test_header_lines_come_first(self):
        ws = self.export(header=self.header)

        self.assertEqual("Форма:", ws["A1"].value)
        self.assertEqual("12-ГА", ws["B1"].value)
        self.assertEqual("Предприятие:", ws["A2"].value)
        self.assertEqual("Тестовая АК", ws["B2"].value)

    def test_period_is_in_the_file(self):
        ws = self.export(header=self.header)

        labels = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                  for r in range(1, len(self.header.lines) + 1)}
        self.assertEqual("Январь 2025 — Март 2025", labels["Период:"])

    def test_sheet_is_no_longer_called_data(self):
        ws = self.export(header=self.header)

        self.assertEqual("12-ГА (свод)", ws.title)

    def test_blank_row_separates_the_header_from_the_table(self):
        ws = self.export(header=self.header)
        separator = len(self.header.lines) + 1

        self.assertIsNone(ws.cell(row=separator, column=1).value)
        self.assertEqual("Показатель", ws.cell(row=separator + 1, column=1).value)

    def test_table_follows_the_header(self):
        ws = self.export(header=self.header)
        first_data_row = len(self.header.lines) + 3

        self.assertEqual("Самолето-километры", ws.cell(row=first_data_row, column=1).value)
        self.assertEqual(1234567, ws.cell(row=first_data_row, column=2).value)

    def test_numbers_below_the_header_are_still_numbers(self):
        """Сдвиг таблицы не должен отменить FUNC-2: ячейка остаётся числовой."""
        ws = self.export(header=self.header)
        first_data_row = len(self.header.lines) + 3

        cell = ws.cell(row=first_data_row, column=2)
        self.assertEqual("n", cell.data_type)
        self.assertEqual("#,##0", cell.number_format)

    def test_form_name_stays_text(self):
        """«12-ГА» не должно превратиться в дату или число."""
        ws = self.export(header=self.header)

        self.assertEqual("s", ws["B1"].data_type)


@unittest.skipUnless(HAS_QT, "PyQt6 не установлен")
class GroupedHeaderShiftsTooTest(WorkbookCase):
    """Двухуровневая шапка колонок съезжает вместе с таблицей, не разъезжаясь."""

    GROUPS = [(1, 2, "2025 год")]

    def setUp(self):
        super().setUp()
        self.header = header_for(stats={"airline_name": "Тестовая АК"})
        self.top = len(self.header.lines) + 2

    def test_group_label_sits_above_the_column_headers(self):
        ws = self.export(header=self.header, header_groups=self.GROUPS)

        self.assertEqual("2025 год", ws.cell(row=self.top, column=2).value)
        self.assertEqual("Январь 2025", ws.cell(row=self.top + 1, column=2).value)

    def test_ungrouped_column_spans_both_header_rows(self):
        ws = self.export(header=self.header, header_groups=self.GROUPS)

        merged = {str(rng) for rng in ws.merged_cells.ranges}
        self.assertIn(f"A{self.top}:A{self.top + 1}", merged)
        self.assertEqual("Показатель", ws.cell(row=self.top, column=1).value)

    def test_data_starts_under_both_header_rows(self):
        ws = self.export(header=self.header, header_groups=self.GROUPS)

        self.assertEqual("Самолето-километры", ws.cell(row=self.top + 2, column=1).value)


@unittest.skipUnless(HAS_QT, "PyQt6 не установлен")
class WithoutHeaderNothingMovesTest(WorkbookCase):
    """Экспорт без шапки остаётся прежним: вызов из чужого кода не ломается."""

    def test_table_still_starts_at_the_first_row(self):
        ws = self.export()

        self.assertEqual("Показатель", ws["A1"].value)
        self.assertEqual("Самолето-километры", ws["A2"].value)
        self.assertEqual("Данные", ws.title)


@unittest.skipUnless(HAS_QT, "PyQt6 не установлен")
class WidgetRemembersWhatItShowsTest(unittest.TestCase):
    """Связка: описать выгрузку может только тот, кто загрузил данные.

    Предприятие и период приходят не из экспорта, а из последнего показанного
    отчёта, — и если виджет перестанет их запоминать, шапка снова окажется
    пустой, а сборка сама по себе останется исправной.
    """

    def make_widget(self):
        from forms.widgets.data_table_widget import DataTableWidget

        widget = DataTableWidget()
        self.addCleanup(widget.deleteLater)
        return widget

    def test_pivot_header_describes_the_last_report(self):
        from unittest.mock import patch

        widget = self.make_widget()
        loaded = {
            "rows": [], "headers": ["Показатель"], "keys": ["indicator"], "groups": [],
            # Набор ключей — как у настоящего свода по одной АК: строка под
            # таблицей читает и месяцы тоже.
            "stats": {"airline_name": "Тестовая АК", "indicators": 31, "months": 3},
        }
        filters = {"period_from": (2025, 1), "period_to": (2025, 3)}

        with patch.object(widget.data_controller, "load_pivot_data", return_value=loaded):
            widget.load_data(MODE_AIRLINE, filters)
        header = widget.export_header(user="ваня")

        self.assertEqual("Тестовая АК", value_of(header, "Предприятие"))
        self.assertEqual("Январь 2025 — Март 2025", value_of(header, "Период"))
        self.assertEqual("31", value_of(header, "Показателей"))
        self.assertEqual("ваня", value_of(header, "Пользователь"))

    def test_header_before_any_load_still_makes_sense(self):
        """Выгрузка пустого экрана не должна падать — и не должна врать."""
        header = self.make_widget().export_header()

        self.assertEqual("12-ГА", value_of(header, "Форма"))
        self.assertIsNone(value_of(header, "Период"))


if __name__ == "__main__":
    unittest.main()

"""Слой `controllers/` не знает ни про интерфейс (ARCH-2), ни про `services/` (ARCH-14).

Экспортёр сам обходил `QTableView` и сам показывал `QMessageBox`: слой, который
по замыслу интерфейса не касается, был кодом интерфейса. Позвать его из теста
или из командной строки было нельзя — модальное окно остановило бы и то и другое,
а об ошибке он сообщал не вызывающему, а пользователю, возвращая наружу `False`
без причины.

Сборка книги теперь принимает готовые заголовки и строки значений и об ошибке
сообщает исключением. Чтение модели и окна сообщений — в `forms/table_export.py`.

Второй запрет — на `services/`. Пакеты были замкнуты в кольцо: `data_controller`
звал `services.airline_ind_service`, а тот звал `controllers.AirlineIndController`.
На уровне модулей это разрешалось порядком импорта, поэтому не мешало ничему и не
падало никогда — но уровни из двух пакетов не строились, и правило «`controllers/`
ниже `services/`» нельзя было ни записать, ни проверить. Обёртки над репозиториями
переехали в `controllers/`, и теперь пакеты не знают друг о друге вовсе.

Проверяется так же, как граница пакета данных в ARCH-3: запретом импорта в
отдельном процессе, а не разглядыванием кода.

Здесь же — проверка обратного свойства: что модуль слоя вообще кому-то нужен
(ARCH-17). Оба вопроса об одном — о рёбрах графа импортов, только первый требует,
чтобы ребра не было, а второй — чтобы хоть одно было.
"""

import os
import re
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook

from controllers.export_controller import ExportController

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Запрет ставится подменой в sys.modules: `import PyQt6` падает и в том случае,
# если пакет установлен, — а он установлен, иначе половина прогона пропускалась бы.
#
# Модули пакета не перечисляются руками, а обходятся `pkgutil`: список, набранный
# вручную, отстаёт от пакета молча, и новый модуль оказался бы вне проверки — то
# есть ровно тот модуль, который её и завалил бы.
_IMPORT_CONTROLLERS_ALONE = """
import importlib
import pkgutil
import sys

for forbidden in ('PyQt6', 'services'):
    sys.modules[forbidden] = None
    try:
        importlib.import_module(forbidden)
    except ImportError:
        pass
    else:
        raise AssertionError(
            'запрет на %s не сработал — проверка ничего не значит' % forbidden
        )

import controllers

# walk_packages, а не iter_modules: с появлением controllers/reports/ (ARCH-15)
# плоский обход поднимал бы сам подпакет и ни одного модуля внутри него.
modules = list(pkgutil.walk_packages(controllers.__path__, prefix='controllers.'))
if len(modules) < 2:
    raise AssertionError('обход нашёл меньше двух модулей — проверять нечего')

for info in modules:
    importlib.import_module(info.name)
"""


class ControllersNeedNoGuiTest(unittest.TestCase):
    def test_layer_imports_without_pyqt_or_services(self):
        """Весь пакет поднимается в процессе, где нет ни PyQt6, ни `services/`."""
        result = subprocess.run(
            [sys.executable, "-c", _IMPORT_CONTROLLERS_ALONE],
            cwd=PROJECT_ROOT, capture_output=True, text=True,
            env={**os.environ, "PYTHONPATH": str(PROJECT_ROOT)},
        )
        self.assertEqual(0, result.returncode, result.stderr)

    def _importers_of(self, package: str):
        """Файлы слоя, импортирующие пакет. rglob: подпакет `reports/` — тоже слой."""
        root = PROJECT_ROOT / "controllers"
        return [
            str(path.relative_to(root)) for path in root.rglob("*.py")
            if re.search(rf"^\s*(import|from)\s+{package}", path.read_text(encoding="utf-8"), re.M)
        ]

    def test_no_source_in_the_layer_mentions_pyqt(self):
        """Проверка на будущее: следующий диалог не должен приехать обратно."""
        self.assertEqual([], self._importers_of("PyQt6"))

    def test_layer_does_not_reach_into_forms(self):
        """Обратная зависимость ушла вместе с чтением модели: роль RAW_VALUE_ROLE — там же, где модели."""
        self.assertEqual([], self._importers_of("forms"))

    def test_layer_does_not_reach_into_services(self):
        """Кольцо ARCH-14: `controllers/` не зовёт `services/` ни одним модулем."""
        self.assertEqual([], self._importers_of("services"))


class NoUnreachableModulesTest(unittest.TestCase):
    """Модуль, которого никто не импортирует, — мёртвый код (ARCH-17).

    `UserController` пролежал так всю жизнь проекта: девять строк, один метод
    `get_user_by_login` и ни одной ссылки — запрос, который он предлагал, к тому
    времени уже стоял внутри `AuthService.sign_in`. Реестр числил его «артефактом
    первоначальной структуры», то есть кодом не на своём месте; на месте
    выяснилось, что места у него нет вовсе.

    `ruff` такого не видит: `F401` — про неиспользованный импорт внутри модуля, а
    не про модуль, который не импортирует никто. Компиляция и прогон тоже молчат —
    мёртвый модуль не мешает ничему, кроме чтения.

    Проверяются `controllers/` и `services/`: точек входа в них нет, каждый модуль
    обязан кем-то вызываться. `forms/`, `main.py` и миграции под правило не
    попадают — их зовёт Qt, ярлык и Alembic соответственно.
    """

    LAYERS = ("controllers", "services")

    def _sources(self):
        for path in PROJECT_ROOT.rglob("*.py"):
            parts = path.relative_to(PROJECT_ROOT).parts
            if any(part.startswith(".") or part in ("build", "dist") for part in parts):
                continue
            yield path

    def test_every_module_of_the_layer_is_imported_by_somebody(self):
        sources = list(self._sources())
        orphans = []

        for layer in self.LAYERS:
            for module in sorted((PROJECT_ROOT / layer).rglob("*.py")):
                if module.name == "__init__.py":
                    continue
                dotted = ".".join(module.relative_to(PROJECT_ROOT).with_suffix("").parts)
                parent, _, leaf = dotted.rpartition(".")
                # Две формы записи: `from пакет.модуль import имя` и
                # `from пакет import модуль` — вторая иначе читалась бы как «никто».
                referenced = re.compile(
                    rf"^\s*(?:from|import)\s+{re.escape(dotted)}\b"
                    rf"|^\s*from\s+{re.escape(parent)}\s+import\s+.*\b{re.escape(leaf)}\b",
                    re.M,
                )
                if not any(
                    referenced.search(other.read_text(encoding="utf-8"))
                    for other in sources
                    if other != module
                ):
                    orphans.append(dotted)

        self.assertEqual([], orphans)


class WorkbookIsWrittenWithoutQtTest(unittest.TestCase):
    """Сборка книги вызывается напрямую — без окна, без модели, без QApplication."""

    def setUp(self):
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        self.path = os.path.join(tmp.name, "export.xlsx")

    def write(self, **kwargs):
        ExportController.write_workbook(
            self.path,
            headers=["Показатель", "Январь 2025"],
            rows=[["Самолето-километры", 1234567.0], ["Налет часов", 0.0]],
            **kwargs,
        )
        return load_workbook(self.path).active

    def test_values_reach_the_file(self):
        ws = self.write()

        self.assertEqual("Показатель", ws["A1"].value)
        self.assertEqual("Самолето-километры", ws["A2"].value)
        self.assertEqual(1234567, ws["B2"].value)

    def test_numbers_stay_numbers(self):
        """FUNC-2 держится и без Qt: тип ячейки задаёт сборщик, а не модель."""
        ws = self.write()

        self.assertEqual("n", ws["B2"].data_type)
        self.assertEqual("#,##0", ws["B2"].number_format)

    def test_failure_is_reported_by_an_exception(self):
        """Прежде метод возвращал False и показывал окно: причина не доходила."""
        with self.assertRaises(OSError):
            ExportController.write_workbook(
                os.path.join(self.path, "нет", "такого", "пути.xlsx"),
                headers=["Показатель"],
                rows=[["Самолето-километры"]],
            )


if __name__ == "__main__":
    unittest.main()

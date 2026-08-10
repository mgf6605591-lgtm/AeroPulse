"""Ворота импорта: период, распознавание формы и отказ вместо молчаливого успеха.

Проверяется класс дефектов, при котором приложение не падает и не жалуется, а тихо
кладёт в базу неверные данные: подставленный период (DATA-2, DATA-3), разбор чужой
формы (DATA-6) и «успешный» импорт нуля строк (DATA-4).

Книги собираются синтетически: реальные отчёты содержат данные предприятий и в
репозиторий не входят. Тесты на настоящих файлах лежат ниже и пропускаются, если
файлов нет.
"""

import os
import tempfile
import unittest
import xml.etree.ElementTree as ET
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from sqlalchemy.orm import sessionmaker

from db.models.entities import Airline, Airport, Locality
from importers.data_importer import DataImporter
from parsers.f15_xlsx_parser import F15XLSXParser
from parsers.xlsx_parser import XLSXParser
from parsers.xml_parser import XMLParser
from services.import_outcome import PeriodRequired
from services.import_service import ImportService
from services.parse_service import ParseService
from tests.support import MigratedDbCase, make_ga12_workbook
from utils.ga12_layout import GA12_ROW_BY_XML_ROW

F15_ROWS = (
    (1, "Международные регулярные"),
    (2, "Международные нерегулярные"),
    (5, "Внутренние регулярные"),
    (6, "Внутренние нерегулярные"),
    (9, "Все прочие операции"),
)


def make_f15_workbook(path, *, period="за февраль 2026 г.", with_values=True):
    """Книга формы 15-ГА. period=None — ячейка периода остаётся пустой."""
    wb = Workbook()
    ws = wb.active
    ws.title = "15-ГА"
    ws.cell(row=1, column=1, value="Наименование аэропорта:  Тестовый аэропорт")
    # Служебная строка нумерации граф — как в настоящем бланке. В её графе 2 стоит
    # число 2, поэтому по одному номеру она неотличима от строки 02 формы.
    for col in range(1, 14):
        ws.cell(row=6, column=col, value=col)
    if period is not None:
        ws.cell(row=7, column=1, value=period)  # iloc (6, 0)
    ws.cell(row=8, column=1, value="Коммерческие перевозки")
    for offset, (number, label) in enumerate(F15_ROWS):
        excel_row = 9 + offset
        ws.cell(row=excel_row, column=1, value=label)
        ws.cell(row=excel_row, column=2, value=number)
        if with_values:
            for col in range(3, 14):  # графы 3…13
                ws.cell(row=excel_row, column=col, value=number)
    wb.save(path)
    return path


class WorkbookCase(unittest.TestCase):
    def setUp(self):
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        self.tmp_dir = tmp.name

    def path(self, name: str) -> str:
        return os.path.join(self.tmp_dir, name)


class PeriodSubstitutionTest(WorkbookCase):
    """Неопределённый период не подменяется «январём 2025»."""

    def test_ga12_without_titul_has_no_period(self):
        result = XLSXParser.parse_file(make_ga12_workbook(self.path("a.xlsx"), titul_period=None))
        self.assertIsNone(result["month"])
        self.assertIsNone(result["year"])

    def test_ga12_reads_period_from_titul(self):
        result = XLSXParser.parse_file(make_ga12_workbook(self.path("b.xlsx")))
        self.assertEqual(result["month"], "January")
        self.assertEqual(result["year"], 2025)

    def test_year_is_not_taken_from_form_header(self):
        """Реквизиты бланка не должны становиться отчётным годом (DATA-3).

        «Приказ Росстата … от 2019 г.» в шапке — ровно то, что прежняя эвристика
        принимала за год отчёта, потому что искала первое `20\\d\\d` в первых
        10×10 ячейках листа.
        """
        path = make_ga12_workbook(self.path("c.xlsx"), titul_period=None)
        from openpyxl import load_workbook

        wb = load_workbook(path)
        wb["ГА12"].cell(row=2, column=2, value="Форма утверждена приказом Росстата от 19.12.2019 № 955")
        wb.save(path)

        result = XLSXParser.parse_file(path)
        self.assertIsNone(result["year"])

    def test_f15_without_period_cell_has_no_period(self):
        result = F15XLSXParser.parse_file(make_f15_workbook(self.path("d.xlsx"), period=None))
        self.assertIsNone(result["month"])
        self.assertIsNone(result["year"])

    def test_f15_reads_period_from_data_sheet(self):
        result = F15XLSXParser.parse_file(make_f15_workbook(self.path("e.xlsx")))
        self.assertEqual(result["month"], "February")
        self.assertEqual(result["year"], 2026)


class F15RowSelectionTest(WorkbookCase):
    """Строка нумерации граф не должна попадать в данные."""

    def test_column_number_row_is_not_a_data_row(self):
        """В её графе 2 стоит 2 — по номеру она неотличима от строки 02 формы.

        Разбиралась как строка 02, и номера граф 3…13 уходили в базу значениями
        показателей; настоящая строка 02 затем часть из них перезаписывала.
        """
        result = F15XLSXParser.parse_file(make_f15_workbook(self.path("a.xlsx")))
        values = {i["indicator_code"]: i["value"] for i in result["indicators"]}
        # Строка 02 в фикстуре заполнена числом 2 по всем графам.
        self.assertEqual(2, values["15ГА-R02-ВС"])
        self.assertEqual(2, values["15ГА-R02-ПЧ_ВСЕГО"])

    def test_no_duplicate_indicator_codes(self):
        result = F15XLSXParser.parse_file(make_f15_workbook(self.path("b.xlsx")))
        codes = [i["indicator_code"] for i in result["indicators"]]
        self.assertEqual(len(codes), len(set(codes)))


class FormDetectionTest(WorkbookCase):
    """Форма определяется содержимым файла, а не выбором пользователя (DATA-6)."""

    def test_ga12_reports_airline_regardless_of_user_choice(self):
        path = make_ga12_workbook(self.path("a.xlsx"))
        result = ParseService.parse_file(path, entity_type="airport", entity_id=1)
        self.assertEqual(result["data_type"], "airline")

    def test_f15_reports_airport_regardless_of_user_choice(self):
        path = make_f15_workbook(self.path("b.xlsx"))
        result = ParseService.parse_file(path, entity_type="airline", entity_id=1)
        self.assertEqual(result["data_type"], "airport")

    def test_forms_are_not_confused_with_each_other(self):
        self.assertFalse(F15XLSXParser.is_f15_workbook(make_ga12_workbook(self.path("c.xlsx"))))
        self.assertTrue(F15XLSXParser.is_f15_workbook(make_f15_workbook(self.path("d.xlsx"))))

    def test_sheet_name_alone_does_not_make_it_a_form(self):
        """Лист с именем «ГА12», но без показателей, бланком не считается."""
        wb = Workbook()
        wb.active.title = "ГА12"
        wb.active.cell(row=1, column=1, value="Пустой лист")
        path = self.path("e.xlsx")
        wb.save(path)
        with self.assertRaises(ValueError):
            XLSXParser.parse_file(path)

    def test_unrecognized_workbook_is_refused(self):
        wb = Workbook()
        wb.active.cell(row=1, column=1, value="Совершенно посторонняя таблица")
        path = self.path("f.xlsx")
        wb.save(path)
        with self.assertRaises(ValueError):
            ParseService.parse_file(path, entity_type="airline", entity_id=1)

    def test_uppercase_extension_is_accepted(self):
        """Расширение сравнивается без учёта регистра (BUG-6).

        Windows регистр в именах не различает, и отчёт, присланный как «.XLSX»,
        открывался у пользователя, но не открывался программой.
        """
        path = make_ga12_workbook(self.path("h.XLSX"))
        result = ParseService.parse_file(path, entity_type="airline", entity_id=1)
        self.assertEqual(result["data_type"], "airline")

    def test_titul_sheet_is_not_parsed_as_report(self):
        """Прежний откат на первый лист книги разбирал «Титул» как отчёт (DATA-4)."""
        path = make_ga12_workbook(self.path("g.xlsx"), sheet_title="Титул", labels=())
        with self.assertRaises(ValueError):
            XLSXParser.parse_file(path)


class ImporterRefusalTest(MigratedDbCase):
    """Импортёр отказывает вместо подстановки умолчаний и мнимого успеха."""

    def setUp(self):
        super().setUp()
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)
        with self.Session() as session:
            session.add(Airline(code="AAA", name="Тестовая АК"))
            session.commit()
            self.airline_id = session.query(Airline).one().id

    def payload(self, indicators, month="March", year=2025):
        return {
            "entity_type": "airline",
            "data_type": "airline",
            "entity_id": self.airline_id,
            "airline": {"name": "Тестовая АК", "code": "AAA", "id": self.airline_id},
            "month": month,
            "year": year,
            "indicators": indicators,
        }

    def do_import(self, indicators, **kwargs):
        with self.Session() as session:
            return DataImporter._import_airline_data(session, self.payload(indicators, **kwargs))

    def indicator_row(self):
        return {
            "indicator_code": "965",
            "indicator_name": "Самолето-километры",
            "measure": "тыс.сам.-км",
            "route_type": "trunk",
            "regularity": "regular",
            "value": 100,
        }

    def test_zero_records_is_a_failure(self):
        result = self.do_import([])
        self.assertFalse(result.success)
        self.assertEqual(result.imported, 0)
        self.assertEqual(result.updated, 0)

    def test_missing_month_is_refused(self):
        result = self.do_import([self.indicator_row()], month=None)
        self.assertFalse(result.success)
        self.assertIn("период", result.message.lower())

    def test_missing_year_is_refused(self):
        result = self.do_import([self.indicator_row()], year=None)
        self.assertFalse(result.success)

    def test_unknown_month_name_is_refused_not_defaulted(self):
        """Нераспознанное название месяца прежде превращалось в Months.January."""
        result = self.do_import([self.indicator_row()], month="Sarlacc")
        self.assertFalse(result.success)

    def test_valid_import_still_succeeds(self):
        result = self.do_import([self.indicator_row()])
        self.assertTrue(result.success)
        self.assertEqual(result.imported, 1)


class ImportServiceGateTest(MigratedDbCase, WorkbookCase):
    """Сквозная проверка: файл без периода не импортируется, а запрашивает период."""

    def setUp(self):
        MigratedDbCase.setUp(self)
        WorkbookCase.setUp(self)
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)
        with self.Session() as session:
            session.add(Airline(code="AAA", name="Тестовая АК"))
            locality = Locality(name="Город")
            session.add(locality)
            session.flush()
            session.add(Airport(code="XXX", name="Аэропорт", locality_id=locality.id))
            session.commit()
            self.airline_id = session.query(Airline).one().id
            self.airport_id = session.query(Airport).one().id
        patcher = patch("services.import_service.get_session", self.Session)
        patcher.start()
        self.addCleanup(patcher.stop)

    def test_file_without_period_asks_instead_of_importing(self):
        path = make_ga12_workbook(self.path("a.xlsx"), titul_period=None)
        result = ImportService.import_file(path, entity_type="airline", entity_id=self.airline_id)
        self.assertIsInstance(result, PeriodRequired)

    def test_period_supplied_by_caller_lets_import_through(self):
        """Именно так mainWin повторяет импорт после ответа пользователя в диалоге."""
        path = make_ga12_workbook(self.path("b.xlsx"), titul_period=None)
        result = ImportService.import_file(
            path, entity_type="airline", entity_id=self.airline_id, month="July", year=2024
        )
        self.assertTrue(result.success, result.message)
        self.assertEqual(result.month, "July")
        self.assertEqual(result.year, 2024)

    def test_ga12_file_is_refused_for_airport(self):
        """Тот самый сценарий DATA-6: бланк авиакомпании при выбранном аэропорте."""
        path = make_ga12_workbook(self.path("c.xlsx"))
        result = ImportService.import_file(path, entity_type="airport", entity_id=self.airport_id)
        self.assertFalse(result.success)
        self.assertIn("12-ГА", result.message)

    def test_f15_file_is_refused_for_airline(self):
        """Обратная сторона той же защиты: бланк аэропорта при выбранной авиакомпании."""
        path = make_f15_workbook(self.path("d.xlsx"))
        result = ImportService.import_file(path, entity_type="airline", entity_id=self.airline_id)
        self.assertFalse(result.success)
        self.assertIn("15-ГА", result.message)


REAL_GA12 = "12-га январь.xlsx"
REAL_F15 = "ФКП АС 15-ГА Февраль 2026 год(1).xlsx"

# XML-выгрузки 12-ГА: три авиакомпании по двенадцать месяцев 2025 года.
REAL_GA12_XML_DIR = "12-ГА 3 авиакомпании"
# Тот же отчёт, что и REAL_GA12: АО «Авиакомпания "АЛРОСА"», январь 2025 года.
# Именно эта пара закрывает сравнение форматов на настоящих файлах (INFRA-1).
REAL_GA12_XML = os.path.join(REAL_GA12_XML_DIR, "0615106_12_12_2269_2025_1.xml")


@unittest.skipUnless(os.path.exists(REAL_GA12), f"нет файла {REAL_GA12}")
class RealGa12FileTest(unittest.TestCase):
    """Настоящий бланк 12-ГА. Файл в репозиторий не входит — тест пропускается без него."""

    def test_recognized_and_period_read(self):
        result = XLSXParser.parse_file(REAL_GA12)
        self.assertEqual(result["data_type"], "airline")
        self.assertEqual(result["sheet_name"], "ГА12")
        self.assertEqual((result["month"], result["year"]), ("January", 2025))
        self.assertTrue(result["indicators"])

    def values(self) -> dict:
        result = XLSXParser.parse_file(REAL_GA12)
        return {(i["indicator_code"], i["route_type"]): i["value"] for i in result["indicators"]}

    def test_values_match_the_blank(self):
        """Сверка с бланком построчно: графа 6 «Внутренние — всего».

        Проверка есть именно потому, что прежний тест ограничивался «показатели
        прочитались». Строки листа адресовались индексами и были смещены на
        единицу: 452 самолёто-километра лежали в базе как «Отправлений воздушных
        судов», 223 отправления — как «Налет часов», 642 часа налёта — как
        «Перевезено пассажиров». Ни одного признака ошибки при этом не было.
        """
        values = self.values()
        self.assertEqual(Decimal("452"), values[("965", "local")])
        self.assertEqual(Decimal("223"), values[("642", "local")])
        self.assertEqual(Decimal("642"), values[("356", "local")])
        self.assertEqual(Decimal("27666"), values[("792", "local")])
        # Сравнение именно с Decimal, а не с 140.62: двоичный float таким числом
        # не является, и равенство здесь не выполнялось бы (BUG-4).
        self.assertEqual(Decimal("140.62"), values[("168", "local")])
        self.assertEqual(Decimal("24.21"), values[("168п", "local")])

    def test_ton_detail_sums_up_to_its_parent(self):
        """а) пассажирский + б) грузовой + в) почтовый = «Выполненный тоннокилометраж».

        Равенство держится на бланке само по себе, поэтому служит проверкой того,
        что все четыре строки опознаны верно, а не только прочитаны.
        """
        values = self.values()
        for route_type in ("local", "interregional"):
            with self.subTest(route_type=route_type):
                detail = sum(values[(code, route_type)] for code in ("450пас", "450гр", "450пч"))
                self.assertAlmostEqual(values[("450", route_type)], detail, places=2)

    def test_non_commercial_section_has_only_the_blank_row(self):
        """В бланке некоммерческих полётов одна строка — «Налет часов»."""
        result = XLSXParser.parse_file(REAL_GA12)
        codes = {i["indicator_code"] for i in result["indicators"] if i["regularity"] == "non_commercial"}
        self.assertEqual({"356нк"}, codes)


@unittest.skipUnless(os.path.exists(REAL_F15), f"нет файла {REAL_F15}")
class RealF15FileTest(unittest.TestCase):
    """Настоящий бланк 15-ГА в XLSX."""

    def test_recognized_and_period_read(self):
        result = F15XLSXParser.parse_file(REAL_F15)
        self.assertEqual(result["data_type"], "airport")
        self.assertEqual(result["sheet_name"], "15-ГА")
        self.assertEqual((result["month"], result["year"]), ("February", 2026))

    def test_values_match_the_blank(self):
        """Строка 05 «Внутренние регулярные» — сверка с бланком по всем графам."""
        result = F15XLSXParser.parse_file(REAL_F15)
        values = {i["indicator_code"]: i["value"] for i in result["indicators"]}
        self.assertEqual(values["15ГА-R05-ВС"], 610)
        self.assertEqual(values["15ГА-R05-ПАС_ОТП"], 13871)
        self.assertEqual(values["15ГА-R05-ПАС_ПРИН"], 15233)
        self.assertEqual(values["15ГА-R05-ПАС_ВСЕГО"], 29104)
        self.assertEqual(values["15ГА-R05-ПАС_ТРАНЗ"], 508)
        # Графа «всего» бланка равна сумме составляющих.
        self.assertAlmostEqual(
            values["15ГА-R05-ГР_ВСЕГО"],
            values["15ГА-R05-ГР_ОТГР"] + values["15ГА-R05-ГР_РАЗГ"],
            places=6,
        )

    def test_row_09_placeholders_are_not_values(self):
        """В строке 09 графы закрыты знаком «Х» — это не ноль и не значение."""
        result = F15XLSXParser.parse_file(REAL_F15)
        codes = {i["indicator_code"] for i in result["indicators"]}
        self.assertNotIn("15ГА-R09-ПАС_ОТП", codes)

    def test_no_duplicate_indicator_codes(self):
        """Дубль означал бы, что в один ключ отчётности легли две строки файла."""
        codes = [i["indicator_code"] for i in F15XLSXParser.parse_file(REAL_F15)["indicators"]]
        self.assertEqual(len(codes), len(set(codes)))

    def test_column_number_row_did_not_leak_into_data(self):
        """В строке 02 бланка заполнены только графы «всего», и там нули."""
        values = {
            i["indicator_code"]: i["value"]
            for i in F15XLSXParser.parse_file(REAL_F15)["indicators"]
        }
        self.assertEqual({0.0}, {v for k, v in values.items() if "-R02-" in k})


@unittest.skipUnless(
    os.path.exists(REAL_GA12) and os.path.exists(REAL_GA12_XML),
    "нет пары «XLSX + XML» одного отчёта 12-ГА",
)
class RealGa12FormatParityTest(unittest.TestCase):
    """Один отчёт в двух форматах — на настоящих файлах, а не на собранных в тесте.

    `FormatParityTest` в `test_ga12_layout.py` сравнивает форматы на книге и XML,
    собранных тут же, и потому проверяет только раскладку: в синтетическом отчёте
    заполнено всё. Настоящая пара добавляет то, чего в ней не воспроизвести, —
    как выгрузка обходится с незаполненным.
    """

    def setUp(self):
        self.from_xlsx = XLSXParser.parse_file(REAL_GA12)
        self.from_xml = XMLParser.parse_file(REAL_GA12_XML)

    @staticmethod
    def by_code(result: dict) -> dict:
        return {(i["indicator_code"], i["route_type"]): i["value"] for i in result["indicators"]}

    def test_both_files_are_the_same_report(self):
        """Сравнивать есть смысл только один отчёт за один период.

        Заодно единственная проверка периода XML на настоящем файле: в выгрузках
        январь записан как `period="1"`, без ведущего нуля.
        """
        self.assertEqual(("January", 2025), (self.from_xlsx["month"], self.from_xlsx["year"]))
        self.assertEqual(("January", 2025), (self.from_xml["month"], self.from_xml["year"]))
        self.assertEqual("airline", self.from_xlsx["data_type"])
        self.assertEqual("airline", self.from_xml["data_type"])

    def test_shared_rows_have_the_same_values(self):
        """Совпадение точное: значения — `Decimal`, и сравниваются как есть."""
        xlsx, xml = self.by_code(self.from_xlsx), self.by_code(self.from_xml)
        shared = set(xlsx) & set(xml)
        self.assertTrue(shared)
        self.assertEqual({key: xlsx[key] for key in shared}, {key: xml[key] for key in shared})

    def test_the_workbook_has_every_row_of_the_xml(self):
        """Строка, которую выгрузка написала, обязана найтись и в книге."""
        self.assertEqual(set(), set(self.by_code(self.from_xml)) - set(self.by_code(self.from_xlsx)))

    def test_rows_only_in_the_workbook_are_zeros(self):
        """Расхождение наборов — целиком нули, и это свойство выгрузки, а не разбора.

        Бланк печатает 0 в международных графах и в строке некоммерческих полётов;
        XML пустую графу и пустую строку просто не пишет. Свести наборы нечем:
        дописать нули в XML значило бы придумать значения, которых в файле нет,
        а пропустить их в XLSX — потерять заявленное организацией «перевозок не
        было». Поэтому проверяется не равенство наборов, а состав разницы.
        """
        xlsx, xml = self.by_code(self.from_xlsx), self.by_code(self.from_xml)
        surplus = {key: value for key, value in xlsx.items() if key not in xml}
        self.assertTrue(surplus, "нулевых граф в книге не нашлось — сравнивать нечего")
        self.assertEqual([], [(key, value) for key, value in surplus.items() if value != 0])

    def test_names_measures_and_sections_agree(self):
        """Один код показателя не должен означать в двух форматах разное."""

        def described(result: dict) -> dict:
            return {
                i["indicator_code"]: (i["indicator_name"], i["measure"], i["regularity"])
                for i in result["indicators"]
            }

        xlsx, xml = described(self.from_xlsx), described(self.from_xml)
        self.assertEqual(
            {code: xlsx[code] for code in xml},
            xml,
        )


@unittest.skipUnless(os.path.isdir(REAL_GA12_XML_DIR), f"нет каталога {REAL_GA12_XML_DIR}")
class RealGa12RoundingTest(unittest.TestCase):
    """Округление: сумма разобранных значений сходится с бланком до последнего знака.

    Графа 9 «ИТОГО гр.4+гр.5+гр.6» в базу не идёт — она производная. Здесь она
    берётся из файла как эталон: сложить то, что разобрал парсер, и получить ровно
    её. Равенство держится на `Decimal`; на `float` оно рвётся в младшем разряде
    (BUG-4), и проверяется весь годовой комплект, а не одна пара файлов.
    """

    @staticmethod
    def files() -> list:
        return sorted(Path(REAL_GA12_XML_DIR).glob("*.xml"))

    @staticmethod
    def control_totals(path: Path) -> dict:
        """Графа 9 по каждой строке — прямо из файла, минуя парсер."""
        root = ET.parse(str(path)).getroot()
        totals = {}
        for row in root.findall(".//sections/section/row"):
            for col in row.findall("col"):
                if col.get("code") == "9" and col.text:
                    totals[int(row.get("code"))] = Decimal(col.text.strip())
        return totals

    @staticmethod
    def by_code(path: Path) -> dict:
        parsed = XMLParser.parse_file(str(path))
        return {(i["indicator_code"], i["route_type"]): i["value"] for i in parsed["indicators"]}

    def test_route_types_add_up_to_the_control_column(self):
        """Международные плюс внутренние — это и есть графа 9.

        Графы 7 и 8 подписаны «из них» и входят в графу 6, поэтому в сумму не
        берутся: сложить все четыре вида сообщения значило бы посчитать внутренние
        дважды.
        """
        for path in self.files():
            values = self.by_code(path)
            for xml_row, total in self.control_totals(path).items():
                row = GA12_ROW_BY_XML_ROW.get(xml_row)
                if row is None:
                    continue
                # Слагаемые складываются с нуля-целого: если значения однажды
                # снова станут float, расхождение будет видно числом, а не
                # ошибкой сложения разных типов.
                parts = sum(values.get((row.code, route), 0) for route in ("trunk", "local"))
                with self.subTest(file=path.name, row=row.code):
                    self.assertEqual(total, parts)

    def test_ton_detail_adds_up_to_its_parent(self):
        """а) пассажирский + б) грузовой + в) почтовый = «Выполненный тоннокилометраж».

        Равенство есть в самом бланке, поэтому расхождение означало бы либо
        потерянную точность, либо строку, опознанную не под своим кодом.
        """
        detail = ("450пас", "450гр", "450пч")
        for path in self.files():
            values = self.by_code(path)
            for route in ("trunk", "local", "interregional", "subsidir"):
                if ("450", route) not in values:
                    continue
                parts = sum(values.get((code, route), 0) for code in detail)
                with self.subTest(file=path.name, route_type=route):
                    self.assertEqual(values[("450", route)], parts)

    def test_the_files_still_hold_sums_that_float_cannot_take(self):
        """Проверки выше имеют смысл, пока такие суммы в комплекте есть.

        Если однажды не останется ни одной, равенство станет выполняться и на
        `float` — обе проверки перестанут ловить BUG-4, ничем этого не показав.
        """
        broken = 0
        for path in self.files():
            root = ET.parse(str(path)).getroot()
            for row in root.findall(".//sections/section/row"):
                cols = {c.get("code"): c.text.strip() for c in row.findall("col") if c.text}
                if "9" not in cols:
                    continue
                parts = [cols[code] for code in ("4", "5", "6") if code in cols]
                exact = sum((Decimal(part) for part in parts), Decimal(0)) == Decimal(cols["9"])
                if exact and sum(float(part) for part in parts) != float(cols["9"]):
                    broken += 1
        self.assertGreater(broken, 0)


if __name__ == "__main__":
    unittest.main()

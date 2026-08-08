"""Выгрузка в XLSX: числа числами, текст текстом (FUNC-2, FUNC-9, BUG-27).

Экспорт брал у модели `DisplayRole` — готовую строку для экрана, — и все числовые
ячейки книги оказывались текстом: Excel помечал их «Число сохранено как текст», и
по выгрузке нельзя было ни суммировать, ни строить диаграммы, ни сортировать
(FUNC-2). Разделитель дробной части при этом оставался точкой, а в русской локали
Excel «1 234.57» не распознаётся как число даже после смены формата ячейки
(BUG-27). Строка, начинающаяся с «=», записывалась формулой (FUNC-9).

Книга собирается настоящим экспортёром и перечитывается `openpyxl`: проверяется
то, что окажется в файле у получателя, а не промежуточные вызовы.
"""

import os
import tempfile
import unittest
from decimal import Decimal

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook

from forms.table_export import export_table_to_excel

try:
    from PyQt6.QtWidgets import QApplication, QTableView
    HAS_QT = True
except ImportError:  # PyQt6 отсутствует — проверки Qt пропускаются
    HAS_QT = False

_app = None


def setUpModule():
    global _app
    if HAS_QT:
        _app = QApplication.instance() or QApplication([])


HEADERS = ["Показатель", "Январь 2025", "Февраль 2025"]
KEYS = ["indicator", "m_2025_January", "m_2025_February"]

# Строка отчёта: целое, дробное, ноль и пустая ячейка — всё, что встречается в своде.
ROWS = [
    {"indicator": "Самолето-километры", "m_2025_January": 1234567.0, "m_2025_February": 1234.567},
    {"indicator": "Налет часов", "m_2025_January": 0.0, "m_2025_February": None},
    # Название приходит из присланного файла, то есть это текст извне.
    {"indicator": "=1+1", "m_2025_January": 5.0, "m_2025_February": 2.5},
]


@unittest.skipUnless(HAS_QT, "PyQt6 не установлен")
class ExportCase(unittest.TestCase):
    def setUp(self):
        from forms.models.pivot_dict_model import PivotDictModel

        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        self.path = os.path.join(tmp.name, "export.xlsx")

        self.model = PivotDictModel()
        self.model.set_source_data(ROWS, HEADERS, KEYS)
        self.view = QTableView()
        self.addCleanup(self.view.deleteLater)
        self.view.setModel(self.model)

    def export(self, header_groups=None):
        self.assertTrue(
            export_table_to_excel(self.view, self.path, header_groups=header_groups)
        )
        return load_workbook(self.path).active


class NumberCellsTest(ExportCase):
    """FUNC-2: числовые ячейки должны быть числами."""

    def test_values_are_written_as_numbers(self):
        ws = self.export()

        self.assertEqual("n", ws["B2"].data_type)
        self.assertEqual(1234567, ws["B2"].value)
        self.assertEqual("n", ws["C2"].data_type)
        self.assertAlmostEqual(1234.567, ws["C2"].value, places=3)

    def test_zero_is_a_number_and_not_an_empty_cell(self):
        ws = self.export()

        self.assertEqual("n", ws["B3"].data_type)
        self.assertEqual(0, ws["B3"].value)

    def test_empty_cell_stays_empty(self):
        """Пустая ячейка читается из книги как None — не как ноль и не как текст."""
        ws = self.export()

        self.assertIsNone(ws["C3"].value)

    def test_digit_grouping_is_left_to_excel(self):
        """Разрядность задаётся форматом ячейки, а не пробелами внутри строки.

        Формат «#,##0.00» — не английская запись: запятая и точка в коде формата
        обозначают разделители, и в русском Excel они выводятся как «1 234,57».
        """
        ws = self.export()

        self.assertEqual("#,##0", ws["B2"].number_format)
        self.assertEqual("#,##0.00", ws["C2"].number_format)

    def test_headers_are_still_written(self):
        ws = self.export()

        self.assertEqual("Показатель", ws["A1"].value)
        self.assertEqual("Январь 2025", ws["B1"].value)


class FormulaInjectionTest(ExportCase):
    """FUNC-9: значение из присланного файла не должно стать формулой."""

    def test_leading_equals_stays_text(self):
        ws = self.export()

        cell = ws["A4"]
        self.assertEqual("s", cell.data_type)
        self.assertEqual("=1+1", cell.value)

    def test_other_formula_starters_stay_text(self):
        from forms.models.pivot_dict_model import PivotDictModel

        rows = [{"indicator": text, "m_2025_January": 1.0, "m_2025_February": 1.0}
                for text in ("+7 (495) 000", "-скидка", "@сотрудник", "=cmd|'/c calc'!A1")]
        self.model = PivotDictModel()
        self.model.set_source_data(rows, HEADERS, KEYS)
        self.view.setModel(self.model)

        ws = self.export()

        for row, text in enumerate(("+7 (495) 000", "-скидка", "@сотрудник", "=cmd|'/c calc'!A1"), start=2):
            with self.subTest(text=text):
                self.assertEqual("s", ws.cell(row=row, column=1).data_type)
                self.assertEqual(text, ws.cell(row=row, column=1).value)

    def test_ordinary_text_is_untouched(self):
        ws = self.export()

        self.assertEqual("Самолето-километры", ws["A2"].value)


class GroupedHeaderExportTest(ExportCase):
    """Разметка с группами месяцев сохраняется — данные съезжают на строку ниже."""

    def test_data_starts_below_two_header_rows(self):
        ws = self.export(header_groups=[(1, 2, "2025 год")])

        self.assertEqual("2025 год", ws["B1"].value)
        self.assertEqual("Январь 2025", ws["B2"].value)
        self.assertEqual(1234567, ws["B3"].value)


@unittest.skipUnless(HAS_QT, "PyQt6 не установлен")
class PivotModelFormattingTest(unittest.TestCase):
    """BUG-27: на экране — русская запись числа."""

    def setUp(self):
        from forms.models.pivot_dict_model import PivotDictModel

        self.model = PivotDictModel()
        self.model.set_source_data(ROWS, HEADERS, KEYS)

    def display(self, row, col):
        from PyQt6.QtCore import Qt

        return self.model.data(self.model.index(row, col), Qt.ItemDataRole.DisplayRole)

    def raw(self, row, col):
        from forms.models.roles import RAW_VALUE_ROLE

        return self.model.data(self.model.index(row, col), RAW_VALUE_ROLE)

    def test_decimal_separator_is_a_comma(self):
        """Прежде здесь была точка: «1 234.57» — чужой формат в русском отчёте."""
        self.assertEqual("1 234,57", self.display(0, 2))

    def test_thousands_are_separated_by_a_space(self):
        self.assertEqual("1 234 567", self.display(0, 1))

    def test_raw_role_returns_the_number_itself(self):
        self.assertEqual(1234.567, self.raw(0, 2))
        self.assertEqual(1234567.0, self.raw(0, 1))

    def test_raw_role_keeps_none_for_empty_cells(self):
        self.assertIsNone(self.raw(1, 2))


@unittest.skipUnless(HAS_QT, "PyQt6 не установлен")
class DetailModelRawValueTest(unittest.TestCase):
    """Подробная таблица: в файл уходят число и подпись перечисления, а не repr."""

    class FakeEnum:
        value = "Январь"

    class FakeRecord:
        def __init__(self):
            self.value = Decimal("140.62")
            self.month = DetailModelRawValueTest.FakeEnum()
            self.year = 2025

    def setUp(self):
        from forms.models.sqlalchemy_table_model import SQLAlchemyTableModel

        self.model = SQLAlchemyTableModel()
        self.model.setHeaders(["Значение", "Месяц", "Год"])
        self.model.setColumnAttributes(["value", "month", "year"])
        self.model.set_source_data([self.FakeRecord()])

    def raw(self, col):
        from forms.models.roles import RAW_VALUE_ROLE

        return self.model.data(self.model.index(0, col), RAW_VALUE_ROLE)

    def test_number_stays_a_number(self):
        self.assertEqual(Decimal("140.62"), self.raw(0))

    def test_enum_becomes_its_caption(self):
        self.assertEqual("Январь", self.raw(1))

    def test_integer_stays_an_integer(self):
        self.assertEqual(2025, self.raw(2))


if __name__ == "__main__":
    unittest.main()

"""Общая обвязка тестов: временная БД, состояния схемы, засев справочников,
синтетические книги бланков.

Каждый тест работает в собственном временном файле. Рабочая БД проекта
(db/database.db) не открывается ни на чтение, ни на запись.
"""

import os
import tempfile
import unittest
from decimal import Decimal
from unittest.mock import patch

from alembic import command
from alembic.script import ScriptDirectory
from openpyxl import Workbook
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import sessionmaker

from controllers.data_controller import DataController
from controllers.reference_cache import reference_cache
from db.database import _sqlite_pragmas
from db.migrator import _config, upgrade_to_head
from db.models.entities import Airline, Indicator
from db.models.enums import Months, RouteType, ShippingRegularity


def db_url(engine) -> str:
    return engine.url.render_as_string(hide_password=False)


def make_engine(path: str):
    """Движок с теми же прагмами, что настраивает приложение.

    Обработчик берётся из db.database, а не переписывается здесь: иначе тесты
    проверяли бы собственную копию настроек, а не то, с чем работает программа
    (в частности foreign_keys=ON).
    """
    engine = create_engine(f"sqlite:///{path}")
    event.listens_for(engine, "connect")(_sqlite_pragmas)
    return engine


def baseline_revision() -> str:
    """Первая ревизия линии — состояние схемы до перехода на Alembic."""
    return ScriptDirectory.from_config(_config("sqlite://")).get_base()


def make_legacy_db(engine, *, without_year: bool = False, without_parent_id: bool = False) -> None:
    """Приводит БД в состояние «до Alembic»: схема baseline без таблицы версий.

    Так выглядела база, созданная прежним Base.metadata.create_all(). Флаги
    воспроизводят ещё более старые установки, где колонок year и parent_id
    не было и init_db() дописывал их вручную.
    """
    command.upgrade(_config(db_url(engine)), baseline_revision())

    # Правка схемы в обход приложения: отдельное подключение без прагм, иначе
    # включённые внешние ключи не дают пересобрать таблицу показателей.
    surgery = create_engine(db_url(engine))
    try:
        with surgery.begin() as conn:
            conn.execute(text("DROP TABLE alembic_version"))
            if without_year:
                conn.execute(text("ALTER TABLE airlineInd DROP COLUMN year"))
                conn.execute(text("ALTER TABLE airportInd DROP COLUMN year"))
            if without_parent_id:
                # Колонка входит в табличный FK на саму себя, DROP COLUMN её не берёт:
                # воспроизводим таблицу в том виде, в каком она была до parent_id.
                conn.execute(text("""
                    CREATE TABLE indicators_without_parent (
                        id INTEGER NOT NULL,
                        name VARCHAR(50) NOT NULL,
                        code VARCHAR(20) NOT NULL,
                        measure VARCHAR(20) NOT NULL,
                        PRIMARY KEY (id),
                        UNIQUE (code),
                        UNIQUE (id)
                    )
                """))
                conn.execute(text(
                    "INSERT INTO indicators_without_parent SELECT id, name, code, measure FROM indicators"
                ))
                conn.execute(text("DROP TABLE indicators"))
                conn.execute(text("ALTER TABLE indicators_without_parent RENAME TO indicators"))
    finally:
        surgery.dispose()


def seed_reference_data(engine) -> None:
    """Минимальный набор справочников: авиакомпания с рейсом, аэропорт, показатель."""
    with engine.begin() as conn:
        conn.execute(text("INSERT INTO airlines (id, code, name) VALUES (1, 'AAA', 'Тестовая АК')"))
        conn.execute(text("INSERT INTO routes (id, type, regularity) VALUES (1, 'trunk', 'regular')"))
        conn.execute(text("INSERT INTO shipping (id, route_id, airline_id) VALUES (1, 1, 1)"))
        conn.execute(text("INSERT INTO airport_localities (id, name) VALUES (1, 'Город')"))
        conn.execute(text("INSERT INTO airports (id, code, name, locality_id) VALUES (1, 'XXX', 'Аэропорт', 1)"))
        conn.execute(text(
            "INSERT INTO indicators (id, name, code, measure) VALUES (1, 'Налет часов', '356', 'час.')"
        ))


def table_ddl(engine, name: str):
    with engine.connect() as conn:
        return conn.execute(
            text("SELECT sql FROM sqlite_master WHERE name = :name"), {"name": name}
        ).scalar()


def scalar(engine, sql: str):
    with engine.connect() as conn:
        return conn.execute(text(sql)).scalar()


# Строки бланка 12-ГА: (№ строки из графы 2, название, код по ОКЕИ).
# Взяты из самого бланка, а не из раскладки парсера: фикстура, собранная по коду
# парсера, подтверждает только саму себя — так и держался незамеченным сдвиг строк
# на единицу, при котором каждое значение уходило в базу под соседним кодом.
GA12_ROWS = (
    (1, "Самолето-километры", "965"),
    (2, "Отправлений воздушных судов", "642"),
    (3, "Налет часов", "356"),
    (4, "Перевезено пассажиров", "792"),
    (7, "Выполненный пассажирооборот", "423"),
    (9, "Выполненный тоннокилометраж", "450"),
)

# Детализация тоннокилометража: в бланке эти строки идут под «в том числе:»
# и своего номера не имеют.
GA12_DETAIL_ROWS = (
    ("а) пассажирский", "450"),
    ("б) грузовой (включая срочный груз)", "450"),
    ("в) почтовый", "450"),
)


def ga12_cell_value(number: int, col: int) -> int:
    """Значение графы: своё число в каждой графе, чтобы перепутанные графы были видны."""
    return number * 10 + (col - 4)


def ga12_total_value(number: int) -> int:
    """Графа 9 бланка — «ИТОГО гр.4+гр.5+гр.6», производная от трёх первых."""
    return sum(ga12_cell_value(number, col) for col in (5, 6, 7))


def make_ga12_workbook(path, *, titul_period="за январь 2025 год", with_values=True,
                       sheet_title="ГА12", labels=GA12_ROWS, details=GA12_DETAIL_ROWS):
    """Книга формы 12-ГА. titul_period=None — лист «Титул» не создаётся вовсе.

    Разметка повторяет бланк: графа 1 — название, графа 2 — № строки, графа 4 —
    код по ОКЕИ, графы 5…9 — данные по видам сообщения. Отдельно воспроизведена
    служебная строка нумерации граф: в её графе 2 стоит число 2, по которому она
    неотличима от строки 02 бланка.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    # Ячейка A1 задаёт начало используемого диапазона: без неё pandas начал бы
    # читать с первой заполненной строки и все индексы разъехались бы.
    ws.cell(row=1, column=1, value="Форма 12-ГА")

    for col in range(1, 4):
        ws.cell(row=10, column=col, value=str(col))
    for col in range(5, 11):
        ws.cell(row=10, column=col, value=str(col - 1))

    excel_row = 11
    if labels:
        ws.cell(row=excel_row, column=1, value="РЕГУЛЯРНЫЕ КОММЕРЧЕСКИЕ ПЕРЕВОЗКИ")
        excel_row += 1

    def write_values(row: int, number: int) -> None:
        if not with_values:
            return
        for col in range(5, 10):  # графы 4…8 бланка: международные, внутренние, местные, субсидируемые
            ws.cell(row=row, column=col, value=ga12_cell_value(number, col))
        # Графа 9 «ИТОГО гр.4+гр.5+гр.6» — в бланке она есть, в базу не идёт.
        ws.cell(row=row, column=10, value=ga12_total_value(number))

    for number, label, okei in labels:
        ws.cell(row=excel_row, column=1, value=label)
        ws.cell(row=excel_row, column=2, value=number)
        ws.cell(row=excel_row, column=4, value=okei)
        write_values(excel_row, number)
        excel_row += 1

        # Детализация идёт сразу под строкой 9 «Выполненный тоннокилометраж».
        if number == 9 and details:
            ws.cell(row=excel_row, column=1, value="    в том числе:")
            excel_row += 1
            for offset, (detail_label, detail_okei) in enumerate(details, start=1):
                ws.cell(row=excel_row, column=1, value=f"            {detail_label}")
                ws.cell(row=excel_row, column=4, value=detail_okei)
                write_values(excel_row, 90 + offset)
                excel_row += 1

    if titul_period is not None:
        titul = wb.create_sheet("Титул")
        titul.cell(row=1, column=1, value="Титульный лист")
        titul.cell(row=13, column=4, value=titul_period)  # D13
    wb.save(path)
    return path

class TempDbCase(unittest.TestCase):
    """Тест с пустой временной БД (файл ещё не создан)."""

    def setUp(self) -> None:
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        self.db_path = os.path.join(tmp.name, "test.db")
        self.engine = make_engine(self.db_path)
        self.addCleanup(self.engine.dispose)
        # Кеш справочников общий на приложение, а база у каждого теста своя:
        # без сброса список из чужой базы дожил бы до следующего теста.
        reference_cache.clear()
        self.addCleanup(reference_cache.clear)


class MigratedDbCase(TempDbCase):
    """Тест с БД, поднятой миграциями до актуальной версии."""

    def setUp(self) -> None:
        super().setUp()
        upgrade_to_head(self.engine)

# Заглушки записей отчётности для построителей свода. Построителям нужны только
# атрибуты показателя, рейса и периода; поднимать ради этого полный граф связей
# ORM значило бы проверять SQLAlchemy вместо логики свода.


class FakeAirline:
    def __init__(self, aid: int, name: str):
        self.id = aid
        self.name = name


class FakeRoute:
    def __init__(self, route_type: str, regularity: str):
        self.type = RouteType[route_type]
        self.regularity = ShippingRegularity[regularity]


class FakeShipping:
    def __init__(self, airline: FakeAirline, route: FakeRoute):
        self.airline = airline
        self.route = route


class FakeIndicator:
    def __init__(self, code: str, name: str, measure: str, iid: int = 1):
        self.id = iid
        self.code = code
        self.name = name
        self.measure = measure


class FakeRecord:
    """Запись отчётности в том виде, в каком её видят построители свода."""

    def __init__(self, code, name, month, year, value, *, measure="тыс.сам.-км",
                 airline=("Тестовая АК", 1), route_type="trunk", regularity="regular"):
        self.indicator = FakeIndicator(code, name, measure)
        self.shipping = FakeShipping(FakeAirline(airline[1], airline[0]),
                                     FakeRoute(route_type, regularity))
        self.month = Months[month]
        self.year = year
        self.value = Decimal(str(value))
        self.airport = None


class FakeAggregateRow:
    """Строка агрегата — то, что отдаёт база после GROUP BY."""

    def __init__(self, regularity, route_type, airline_id, airline_name,
                 indicator_id, indicator_code, indicator_name, measure,
                 year, month, total, records):
        self.regularity = regularity
        self.route_type = route_type
        self.airline_id = airline_id
        self.airline_name = airline_name
        self.indicator_id = indicator_id
        self.indicator_code = indicator_code
        self.indicator_name = indicator_name
        self.measure = measure
        self.year = year
        self.month = month
        self.total = total
        self.records = records


def aggregate_rows(records):
    """Свернуть заглушки фактов так, как их сгруппировала бы база.

    Тесты по-прежнему описывают отдельные записи отчётности — это понятнее, — а
    построители свода получают агрегат, потому что после PERF-2 читают именно его.
    """
    buckets = {}
    for rec in records:
        route = rec.shipping.route
        airline = rec.shipping.airline
        indicator = rec.indicator
        key = (
            route.regularity, route.type, airline.id, airline.name,
            indicator.id, indicator.code, indicator.name, indicator.measure,
            rec.year, rec.month,
        )
        bucket = buckets.setdefault(key, [0.0, 0])
        bucket[0] += float(rec.value)
        bucket[1] += 1
    return [FakeAggregateRow(*key, total=total, records=count)
            for key, (total, count) in buckets.items()]


class PivotCase(MigratedDbCase):
    """Построители свода на подменённой выборке и настоящем справочнике показателей."""

    def setUp(self):
        super().setUp()
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)
        with self.Session() as session:
            session.add(Indicator(name="Самолето-километры", code="965", measure="тыс.сам.-км"))
            session.add(Indicator(name="Налет часов", code="356", measure="час."))
            # Свод по одной АК читает её название из БД по id.
            session.add(Airline(id=1, code="AAA", name="Тестовая АК"))
            session.commit()

        session_patch = patch("controllers.data_controller.get_session", self.Session)
        session_patch.start()
        self.addCleanup(session_patch.stop)

        self.controller = DataController()

    def with_records(self, records):
        """Подменяет источник свода: службы отдают агрегат, а не сами факты."""
        return (
            patch("controllers.data_controller.AirlineIndicatorService.aggregate",
                  return_value=aggregate_rows(records)),
        )

    def build_all_airlines(self, records):
        (agg,) = self.with_records(records)
        with agg:
            return self.controller._load_pivot_all_airlines({"any": "filter"})

    def build_per_airline_summary(self, records):
        (agg,) = self.with_records(records)
        with agg:
            return self.controller._load_pivot_per_airline_summary({}, airline_id=1)

    def build_per_airline_by_routes(self, records, filters=None):
        (agg,) = self.with_records(records)
        with agg:
            return self.controller._load_pivot_per_airline(filters or {}, airline_id=1)

    def build_multi_airline_by_routes(self, records, filters=None):
        (agg,) = self.with_records(records)
        with agg:
            return self.controller._load_pivot_multi_airline_by_routes(filters or {"any": "filter"})

    @staticmethod
    def row_for_code(result, code):
        for row in result["rows"]:
            if row.get("code") == code:
                return row
        raise AssertionError(f"в своде нет строки с кодом {code}")
